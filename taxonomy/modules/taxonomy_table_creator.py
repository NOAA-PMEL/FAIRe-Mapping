from pathlib import Path
from abc import ABC, abstractmethod
import pandas as pd
import hashlib
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# TODO: Columns left to add :
    # scientificNameAuthorship
    # taxonID
    # taxonID_db (either NCBI, Bold, or PR2, or WORMS, or GBIF)
# TODO: Add assay name to sheet name when saving to the FAIRe excel file???
# TODO: THis only works for Revamp - will need to adjust/expand for other taxonomy methods

class TaxonomyTableCreator(ABC):
    # FAIRE Taxonomy fields (Includes ones we are adding as User Defined fields like Supergroup)
    SEQ_ID = "seq_id"
    DNA_SEQUENCE = "dna_sequence"
    KINGDOM = "kingdom"
    SUPERGROUP = "supergroup" # pr2 sckit
    DIVISION = "division" # pr2 sckit
    SUBDIVISION = "subdivision" # pr2 scikit
    PHYLUM = "phylum"
    CLASS = "class"
    ORDER = "order"
    FAMILY = "family"
    GENUS = "genus"
    SPECIES = "species"
    SPECIFIC_EPITHET = "specificEpithet"
    SCIENTIFIC_NAME = "scientificName"
    SCIENTIFIC_NAME_AUTHORSHIP = "scientificNameAuthorship"
    TAXON_RANK = "taxonRank"
    TAXON_ID_DB = "taxonID_db"
    ACCESSION_ID = "accession_id"
    ACCESSION_ID_REF_DB = "accession_id_ref_db"
    VERBATIM_IDENTFICATION = "verbatimIdentification"
    PERCENT_MATCH = "percent_match"
    PERCENT_QUERY_COVER = "percent_query_cover"
    CONFIDENCE_SCORE = "confidence_score"

    # Other
    TAXA_FAIRE_TEMPLATE_PATH = "/home/poseidon/zalmanek/FAIRe-Mapping/taxonomy/taxa_faire_template.xlsx"

    def __init__(self, taxon_tble_file: str, asv_dna_seq_file: str, final_faire_template_path: str):
        """
        taxon_table_file: The path to the taxonomy table.txt file
        asv_dna_seq_file: The path to the file with the asv/dna sequences
        """
        self.taxon_table_path = Path(taxon_tble_file)
        self.asv_dna_seq_file = Path(asv_dna_seq_file)
        self.final_faire_excel_path = final_faire_template_path
        self.hash_dna_seq_dict = self._create_asv_hash_dict()
    
    @property
    @abstractmethod
    def taxonomy_cols(self):
        """
        Sublasses must define their taxonomy columns
        """
        pass

    @property
    @abstractmethod
    def taxonomy_unknowns(self):
        """
        Subclasses must define their taxonomy unknowns
        """
        pass

    @abstractmethod
    def build_faire_df(self) -> pd.DataFrame:
        """
        Create the final FAIRe Taxonomy data frame. An abstract method
        because child classes based on bioinformatics methods should have
        this defined differently.
        """
        pass
    
    def _load_tab_text_file_as_df(self, file_path: Path) -> pd.DataFrame:
        """
        Loads a text file that is tab delimated as a pandas data frame.
        """
        df = pd.read_csv(file_path, sep='\t')

        return df
    
    def _create_asv_hash_dict(self) -> dict:
        # creates a dictionary like {'ASV1': {'seq': ACGT, 'hash': 'dafadsf'}}
        # replace this in experiment_run_mapper if end up using the same dictionary.
        asv_hash_dict = {}
        current_asv = None
        seq_lines = []

        print(f"Creating ASV hash dict for {self.asv_dna_seq_file}")
        with open(self.asv_dna_seq_file, 'r') as f:
            for line in f:
                line = line.strip()

                if not line: # Skip empty lines
                    continue
                
                # Get the first ASV
                if current_asv == None:
                    if line.startswith('>'):
                        current_asv = line.replace('>', '')
                        seq_lines = []
                    else:
                        seq_lines.append(line)

                # For other ASVs that aren't first row
                else:
                    if line.startswith('>'):
                        # Add the old current asv first
                        seq = "".join(seq_lines)
                        seq_hash = hashlib.md5(seq.encode()).hexdigest()
                        asv_hash_dict[current_asv] = {
                            'seq': seq, 
                            'hash': seq_hash
                            }

                        # Then create new current_asv and reset seq lines
                        current_asv = line.replace('>', '')
                        seq_lines = []
                    else:
                        seq_lines.append(line)
        
        # *** CRITICAL FIX: Process the last ASV after the loop finishes ***
        if current_asv is not None and seq_lines:
            seq = "".join(seq_lines)
            seq_hash = hashlib.md5(seq.encode()).hexdigest()
            asv_hash_dict[current_asv] = {
                'seq': seq, 
                'hash': seq_hash
                }

        return asv_hash_dict

    def _switch_asv_for_hashes(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Switches the ASV names in column one of the taxonomy df for the 
        corresponding hashes
        """
        # Create flat mapping dictionary of {'ASV_1': 'kasdhfkajdsf'}
        hash_map = {
            seq_id: data['hash'] for seq_id, data in self.hash_dna_seq_dict.items()
        }
        # Replace ASV strings with hashes
        df[self.SEQ_ID] = df[self.SEQ_ID].map(hash_map)
        return df
    
    def _create_dna_sequence_col(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Creates a column called dna_sequence that adds the corresponding
        DNA sequence to based on the hash. Needs to come after the ASVs have
        been switched to hashes.
        """
        hash_map = {
            data['hash']: data['seq'] for data in self.hash_dna_seq_dict.values()
        }
        df[self.DNA_SEQUENCE] = df[self.SEQ_ID].map(hash_map)

        return df

    def _get_taxon_rank(self, row: pd.Series) -> str:
        """
        Gets the most specific column name for the most specific
        name in the Scientific name.
        """
        # Iterate through the taxonomic cols in reverse order (most specific first)
        for col in reversed(self.taxonomy_cols):
            if pd.notna(row[col]) and row[col] not in self.taxonomy_unknowns:
                if col == self.SPECIFIC_EPITHET:
                    return "species"
                else:
                    return col
   
        return "missing: not provided"
    
    def _get_scientific_name(self, row: pd.Series) -> str:
        """
        Gets the most scientific name for the most specific 
        value in the taxonomies.
        """
        for col in reversed(self.taxonomy_cols):
            if pd.notna(row[col]) and row[col] not in self.taxonomy_unknowns:
                return row[col]
        
        return "not applicable"

    def add_final_df_to_FAIRe_excel(self, 
                                    data_df: pd.DataFrame,
                                    template_path: str,
                                    output_path: str,
                                    sheet_name: str = "taxaRaw",
                                    header_row: int = 3):
        
        # Step 1: Read template column names (preserve order)
        template_df = pd.read_excel(
            template_path,
            sheet_name=sheet_name,
            header=header_row - 1, # Convert to 0-indexed
            nrows=0 # Just to get headers
        )
        template_columns = list(template_df.columns)

        # Step 2. Get data columns
        data_columns = set(data_df.columns)

        # Step 3. Find matching columns
        matching_cols = [col for col in template_columns if col in data_columns]
        missing_in_data = [col for col in template_columns if col not in data_columns]
        extra_in_data = [col for col in data_columns if col not in template_columns]

        print(f"Template columns: {len(template_columns)}")
        print(f"Data columns: {len(data_columns)}")
        print(f"Matching columns: {len(matching_cols)}")

        if missing_in_data:
            print(f"\nColumns in template but not in data (will be empty):")
            for col in missing_in_data:
                print(f"    - {col}")

        if extra_in_data:
            print(f"Columns in data but not in template (will be added as User defined fields):")
            for col in extra_in_data:
                print(f"    - {col}")

        # Step 4: Load template workbook (preserve formatting)
        wb = load_workbook(template_path)
        ws = wb[sheet_name]

        # Step 5: Build column index mapping (template column name -> Excel column index)
        col_mapping = {}
        for col_idx, col_name in enumerate(template_columns, start=1):
            if col_name in data_columns:
                col_mapping[col_name] = col_idx

        # Step 6: Add new columns from data that aren't in template
        next_col_idx = len(template_columns) + 1
        new_col_mapping = {}

        if extra_in_data:
            print(f"\nAdding {len(extra_in_data)} new columns to template:")
            for col_name in extra_in_data:
                # Add column header (row 3)
                ws.cell(row=header_row, column=next_col_idx, value=col_name)
                # Add "User defined" in row 2
                ws.cell(row=header_row-1, column=next_col_idx, value="User defined")
                new_col_mapping[col_name] = next_col_idx
                print(f"    - {col_name} (column {next_col_idx})")
                next_col_idx += 1

        # Combine all column mappings
        all_col_ampping = {**col_mapping, **new_col_mapping}

        # Step 7: Write data efficiently
        data_start_row = header_row + 1 # Data starts after header

        print(f"\nWriting {len(data_df)} rows to template . . .")

        # For large datasets, iterate row by row for memory efficiecy
        for row_idx, (_, row) in enumerate(data_df.iterrows()):
            excel_row = data_start_row + row_idx

            for col_name, col_idx in all_col_ampping.items():
                value = row[col_name]

                # Handle NaN/None values
                if pd.isna(value):
                    continue
                ws.cell(row=excel_row, column=col_idx, value=value)

            # Progess indicator for large files
            if (row_idx + 1) % 10000 == 0:
                print(f"    Proceed {row_idx + 1:,} rows . . .")

        # Step 8: save output
        wb.save(output_path)
        print(f"\nSave to: {output_path}")
        print(f"Total rows written: {len(data_df):,}")
                    
                    
                   