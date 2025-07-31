from .faire_mapper import OmeFaireMapper
from .analysis_metadata_mapper import AnalysisMetadataMapper
from .lists import marker_shorthand_to_pos_cont_gblcok_name, project_pcr_library_prep_mapping_dict
from datetime import date, datetime
import pandas as pd
import requests
import base64
import tempfile
import openpyxl
import os
import hashlib
from astral import LocationInfo
from astral.sun import sun
import pytz


# TODO: add project_id to process_whole_project_and_save_to_excel when calling process_analysis_metadata when extracted from projectMetadata input
# TODO: add assay_name to sampleMetadata
# TODO: Find what happend to the Mid.NC.SKQ21 sample (in SampleMetadataMapper)
# TODO: Check 'Arctic Ocean' geo_loc for E1082.SKQ2021, E1083.SKQ2021, E1084.SKQ2021 and maybe other values
# TODO: Add assay1, assay2, etc. to column headers in projectMetadata

class ProjectMapper(OmeFaireMapper):
    
    faire_sample_metadata_sheet_name = "sampleMetadata"
    faire_experiment_run_metadata_sheet_name = "experimentRunMetadata"
    faire_project_metadata_sheet_name = "projectMetadata"

    faire_sample_name_col = "samp_name"
    faire_rel_cont_id_col_name = "rel_cont_id"
    faire_pos_cont_type_col = 'pos_cont_type'
    faire_sample_category_col_name = "samp_category"
    faire_sample_composed_of_col_name = "sample_composed_of"
    faire_assay_name_col = 'assay_name'
    faire_latitude_col = 'decimalLatitude'
    faire_longitude_col = 'decimalLongitude'
    faire_eventDate_col = 'eventDate'
    faire_sunrise_col = 'sunrise_time_utc'
    faire_sunset_col = 'sunset_time_utc'
    faire_sunset_sunrise_method_col = 'sunset_sunrise_method'
    faire_input_read_count_col = 'input_read_count'
    faire_output_read_count_col = 'output_read_count'
    faire_biological_rep_relation_col = 'biological_rep_relation'
    faire_stations_in_5km_col = "station_ids_within_5km_of_lat_lon"
    faire_measurements_from_col = "measurements_from"
    faire_seq_run_id_col = "seq_run_id"
    project_sheet_term_name_col_num = 3
    project_sheet_assay_start_col_num = 5
    project_sheet_project_level_col_num = 4
    
    def __init__(self, config_yaml: str, gh_token: str):

        super().__init__(config_yaml)

        self.config_yaml = config_yaml
        self.gh_token = gh_token
        self.config_file = self.load_config(config_yaml)
        self.project_name = self.config_file['project_name'] if 'project_name' in self.config_file else None # None for NCBI
        self.project_info_google_sheet_id = self.config_file['project_info_google_sheet_id'] if 'project_info_google_sheet_id' in self.config_file else None # None for NCBI
        self.project_info_df = self.load_google_sheet_as_df(google_sheet_id=self.project_info_google_sheet_id, sheet_name='Sheet1', header=0) if 'project_info_google_sheet_id' in self.config_file else None # None for NCBI
        self.project_id = dict(zip(self.project_info_df['faire_field'], self.project_info_df['value']))['project_id'] if 'project_info_google_sheet_id' in self.config_file else None # None for NCBI
        self.mismatch_samp_names_dict = self.config_file['mismatch_sample_names'] if 'mismatch_sample_names' in self.config_file else {}
        self.pooled_samps_dict = self.config_file['pooled_samps'] if 'pooled_samps' in self.config_file else {}
        self.bioinformatics_bebop_path = self.config_file['bioinformatics_bebop_path'] if 'bioinformatics_bebop_path' in self.config_file else None # None if NCBI
        self.bebop_config_file_google_sheet_id = self.config_file['bebop_config_file_google_sheet_id'] if 'bebop_config_file_google_sheet_id' in self.config_file else None # None if NCBI
        self.bioinformatics_software_name = self.config_file['bioinformatics_software_name'] if 'bioinformatics_software_name' in self.config_file else None # None if NCBI
        self.bebop_config_run_col_name = self.config_file['bebop_config_run_col_name'] if 'bebop_config_run_col_name' in self.config_file else None # None if NCBI
        self.bebop_config_marker_col_name = self.config_file['bebop_config_marker_col_name'] if 'bebop_config_marker_col_name' in self.config_file else None # None if NCBI
        self.logging_directory = self.config_file['logging_directory']

        self.pcr_library_dict = {}

    def process_whole_project_and_save_to_excel(self):

        sample_metadata_df, experiment_run_metadata_df = self.process_sample_run_data()

        # Drop submission_type for experiment_run_metadata_df - only applicable to ncbi and needs to be dropped here
        if 'submission_type' in experiment_run_metadata_df.columns:
            experiment_run_metadata_df.drop(columns=['submission_type'])

        data_dir = os.path.dirname(self.final_faire_template_path)

        # Save sample metadata first to excel file and csv, and then use that excel file and save experimentRunMetadata df
        self.add_final_df_to_FAIRe_excel(excel_file_to_read_from=self.faire_template_file, sheet_name=self.faire_sample_metadata_sheet_name, faire_template_df=sample_metadata_df)
        self.save_final_df_as_csv(final_df=sample_metadata_df, sheet_name=self.faire_sample_metadata_sheet_name, header=2, csv_path=f"{data_dir}/sampleMetadata_{self.project_id}.csv")

        self.add_final_df_to_FAIRe_excel(excel_file_to_read_from=self.final_faire_template_path, sheet_name=self.faire_experiment_run_metadata_sheet_name, faire_template_df=experiment_run_metadata_df)
        self.save_final_df_as_csv(final_df=experiment_run_metadata_df, sheet_name=self.faire_experiment_run_metadata_sheet_name, header=2, csv_path=f"{data_dir}/experimentRunMetadata_{self.project_id}.csv")

        # Add projectMetadata, first project_level metadata then assay level metadata
        self.load_project_level_metadata_to_excel_and_save_as_csv()
        self.load_assay_level_metadata_to_excel(final_exp_run_df=experiment_run_metadata_df)
       
        print(f"Excel file saved to {self.final_faire_template_path}")

        # Add analysisMetadata
        # self.process_analysis_metadata(project_id='EcoFOCI_eDNA_2020-23', final_exp_run_df=experiment_run_metadata_df)
    
    def process_sample_run_data(self):
        # Process all csv sets defined in the config file.
        # 1. Combine all sample metadata spreadsheets into one and combine all experiment run metadata spreadsheets into one

        combined_sample_metadata_df = self.create_sample_metadata_df()
        combined_exp_run_df = self.create_exp_run_metadata_df() 

        # Update sample names to match for sample name and rel_cont_id
        combined_sample_metadata_df[self.faire_sample_name_col] = combined_sample_metadata_df[self.faire_sample_name_col].apply(self.update_mismatch_sample_names)
        combined_sample_metadata_df[self.faire_rel_cont_id_col_name] = combined_sample_metadata_df[self.faire_rel_cont_id_col_name].apply(self.update_rel_cont_id_for_mismatch_samps)
        
        # Add in pooled samples and add to rel_cont_id
        sample_metadata_with_pooled = self.add_pooled_samps_to_sample_metadata(sample_df=combined_sample_metadata_df, experiment_run_df=combined_exp_run_df)
        sample_metadata_with_pooled_rel_cont_id = self.update_rel_cont_id_with_pool_samp(sample_df=sample_metadata_with_pooled)

        # Add positive samples to rel_cont_id and sample_df
        pos_samp_map = self.create_positive_sample_map(run_df=combined_exp_run_df, sample_df=sample_metadata_with_pooled_rel_cont_id)
        sample_metadata_with_pooled_rel_cont_id[self.faire_rel_cont_id_col_name] = sample_metadata_with_pooled_rel_cont_id.apply(
            lambda row: self.add_pos_samps_to_rel_cont_id(metadata_row=row, positive_sample_map=pos_samp_map),
            axis=1
        )
        updated_pos_samp_df = self.add_positive_samps_to_samp_df(samp_df=sample_metadata_with_pooled, positive_sample_map=pos_samp_map)

        # Add PCR samples to sample_composed_of
        pcr_updated_df = self.add_pcr_rows_to_samp_df(sample_df=updated_pos_samp_df, exp_run_df=combined_exp_run_df)
    
        # Filter primary dataframe based on samp_name values in sequencing run dataframe
        filtered_samp_df, missing_samples = self._filter_samp_df_by_samp_name(samp_df=pcr_updated_df, associated_seq_df=combined_exp_run_df)
        if missing_samples:
            missing_file_name = 'samps_not_sequenced.csv'
            print(f"\033[35mThere are samples that do not appear to have been sequenced - they don't exist in the experimentRunMetadata, please see {self.logging_directory+missing_file_name} for a list.\33[0m]")
            self.save_list_to_csv(the_list=missing_samples, file_name=missing_file_name)

        # Fill empty values for POSITIVE and pool samples with "not applicable: control sample"
        final_sample_metadata_df = self.fill_empty_vals(df=filtered_samp_df)

        # Filter experiment run metadata to drop rows with sample names missing from SampleMetadata
        final_exp_df, dropped_exp_run_samples = self.filter_exp_run_metadata(final_sample_df=final_sample_metadata_df, exp_run_df=combined_exp_run_df)
        if len(dropped_exp_run_samples) > 0:
            dropped_file_name = 'unrelated_project_samps_dropped.csv'
            print(f"\033[35mThere are samples that were dropped from the experiment run metadata. Please check the {self.logging_directory+dropped_file_name} for a list - make sure they are not part of this project.\33[0m]")
            self.save_list_to_csv(the_list=dropped_exp_run_samples, file_name=dropped_file_name)
            
        # Add assay name to final_sample_df
        final_sample_df = self.add_assay_name_to_samp_df(samp_df=final_sample_metadata_df, exp_run_df=final_exp_df)

        # Add calculated columns to sample df (like sunset, sunrise)
        final_sample_df_with_calc_cols = self.add_post_sample_metadata_calculated_cols(df=final_sample_df)

        # Drop any samples from rel_cont_id or biological_rep_relation that don't exist anymore in the dataframe - may have been dropped because they weren't sequenced.
        final_final_df = self.drop_samps_from_faire_rel_column_that_dropped(df = final_sample_df_with_calc_cols)
        last_df = self.drop_samp_cols_not_meant_for_submission(sample_df=final_final_df)

        return last_df, final_exp_df

    def process_analysis_metadata(self, project_id: str, final_exp_run_df: pd.DataFrame): 
        # Process analysis metadata using AnalyisMetadata class
        analysis_creator = AnalysisMetadataMapper(config_yaml=self.config_yaml,
                                                  bioinformatics_bebop_path=self.bioinformatics_bebop_path,
                                                  bioinformatics_config_google_sheet_id=self.bebop_config_file_google_sheet_id,
                                                  experiment_run_metadata_df=final_exp_run_df,
                                                  bioinformatics_software_name=self.bioinformatics_software_name,
                                                  bebop_config_run_col_name = self.bebop_config_run_col_name,
                                                  bebop_config_marker_col_name = self.bebop_config_marker_col_name,
                                                  project_id=project_id,
                                                  )  
        analysis_creator.process_analysis_metadata()
    
    def create_sample_metadata_df(self) -> pd.DataFrame:
   
        sample_metadata_datasets = self.config_file['datasets']['sample_metadata_csvs']

        # combine sample metadata from specified cruises in config
        samp_metadata_dfs = []
        for samp_metadata in sample_metadata_datasets:
            
            cruise_sample_csv_path = samp_metadata.get('sample_metadata_csv_path')

            cruise_samp_df = self.load_csv_as_df(file_path=cruise_sample_csv_path)
            samp_metadata_dfs.append(cruise_samp_df)

        combined_sample_metadata_df = pd.DataFrame()
        combined_sample_metadata_df = pd.concat(samp_metadata_dfs, ignore_index=True)

        return combined_sample_metadata_df
   
    def create_exp_run_metadata_df(self) -> pd.DataFrame:

        associated_exp_run_datasets = self.config_file['datasets']['associated_sequencing_csvs']

        # combine experimentRunMetadata from specific cruises
        associated_exp_run_dfs = []
        for exp_run in associated_exp_run_datasets:

            exp_run_csv_path = exp_run.get('sequence_run_csv_path')
            exp_run_df = self.load_csv_as_df(file_path=exp_run_csv_path)
            associated_exp_run_dfs.append(exp_run_df)
        
        combined_exp_run_df = pd.DataFrame()
        combined_exp_run_df = pd.concat(associated_exp_run_dfs, ignore_index=True)

        exp_run_df_with_merged_counts = self.update_output_input_counts_for_merged_runs(exp_df = combined_exp_run_df)

        return exp_run_df_with_merged_counts
    
    def update_output_input_counts_for_merged_runs(self, exp_df: pd.DataFrame) -> pd.DataFrame:
        # For OSU/Run3 where there will be the same sample_name and assay_name (part of assay name since different for runs)
        exp_df['assay_group'] = exp_df[self.faire_assay_name_col].str.split('_', n=1).str[0]
        exp_df[self.faire_input_read_count_col] = exp_df.groupby([self.faire_sample_name_col, 'assay_group'])[self.faire_input_read_count_col].transform('sum')
        exp_df[self.faire_output_read_count_col] = exp_df.groupby([self.faire_sample_name_col, 'assay_group'])[self.faire_output_read_count_col].transform('sum')

        # Drop assay_group
        exp_df = exp_df.drop(columns=['assay_group'])

        return exp_df
    
    def add_post_sample_metadata_calculated_cols(self, df: pd.DataFrame) -> pd.DataFrame:
        # Adds columns to sample metadata that are calculated from other columns. (e.g. sunset, sunrise)
        
        # Add sunrise and sunset utc times
        sun_info = df.apply(lambda row: self.get_sun_times_from_iso(metadata_row=row), axis=1, result_type = 'expand')
        df[self.faire_sunrise_col] = sun_info[0]
        df[self.faire_sunset_col] = sun_info[1]
        df[self.faire_sunset_sunrise_method_col] = sun_info[2]

        return df

    def update_mismatch_sample_names(self, sample_name: str) -> str:
        # Update sample metadata sample name to be the sample name that needs to match in experiment run metadata
        if sample_name in self.mismatch_samp_names_dict:
            sample_name = self.mismatch_samp_names_dict.get(sample_name)
        return sample_name

    def update_rel_cont_id_for_mismatch_samps(self, rel_cont_id: str) -> str:
        # Update the rel_cont_id to samples that had mistmatch names
        if pd.notna(rel_cont_id):
            rel_cont_id_list = rel_cont_id.split(' | ')
            updated_rel_cont_id_list = [self.mismatch_samp_names_dict.get(samp) if samp in self.mismatch_samp_names_dict else samp for samp in rel_cont_id_list]
            return ' | '.join(updated_rel_cont_id_list)

    def add_pooled_samps_to_sample_metadata(self, sample_df: pd.DataFrame, experiment_run_df: pd.DataFrame) -> pd.DataFrame:
        # add pooled samples that are relevant to sample metadata spreadsheet
        all_valid_seq_samps = set(experiment_run_df[self.faire_sample_name_col].unique())
        all_valid_sample_samps = set(sample_df[self.faire_sample_name_col].unique())

        # Check if sample name is in pooled sample dictionary and if is add pooled sample to rel_cont_id
        new_rows = []
        for pooled_info in self.pooled_samps_dict:
            pooled_samp = pooled_info['pooled_samp_name']
            for samp in pooled_info['samps_that_were_pooled']:
                if samp in all_valid_sample_samps and pooled_samp in all_valid_seq_samps:
                    sample_category = pooled_info['sample_category']
                    if sample_category == 'negative control':
                        dict_key = 'neg_cont_type'
                    elif sample_category == 'positive control':
                        dict_key = 'pos_cont_type'
                    elif sample_category == 'sample':
                        dict_key = None
                    else:
                        raise ValueError(f"Don't have functionality for this sample category {sample_category} - please see code and possibly update")
                
                    # Create new row for pooled sample - add neg_cont_type or pos_conty_type
                    new_pooled_samp_row = {
                        self.faire_sample_name_col: pooled_samp,
                        self.faire_sample_category_col_name: sample_category,
                        self.faire_sample_composed_of_col_name: ' | '.join([samp for samp in pooled_info['samps_that_were_pooled'] if samp in all_valid_sample_samps])
                    }
                    if dict_key:
                        new_pooled_samp_row[dict_key] = pooled_info['cont_type']
                        new_rows.append(new_pooled_samp_row)

                    # add pooled sample name to subsmaple's rel_cont_id
                    rel_cont_id = sample_df.loc[sample_df[self.faire_sample_name_col] == samp, self.faire_rel_cont_id_col_name].values
                    if ' | ' in rel_cont_id:
                        rel_cont_id = rel_cont_id.split(' | ')
                    else:
                        rel_cont_id = [rel_cont_id]
                    rel_cont_id.append(pooled_samp)
                    # filter to remove not applicable if it existed before
                    if len(rel_cont_id) > 1:
                        filtered_rel_cont_id_list = [id for id in rel_cont_id if "not applicable" not in str(id).lower()]
                    else:
                        filtered_rel_cont_id_list = rel_cont_id
                    sample_df.loc[sample_df[self.faire_sample_name_col] == samp, self.faire_rel_cont_id_col_name] = ' | '.join(filtered_rel_cont_id_list)
                else:
                    print(f"sample {samp} exists in a pooled set according to the config, but may not exist in the sample metadata or the pooled sample may not exist in the experiment run metadata. Please double check this! But moving on without adding it.")
                    
        # Add new pooled samples to sample_df
        if new_rows:
            new_pooled_rows_df = pd.DataFrame(new_rows)
            new_pooled_rows_df = new_pooled_rows_df.drop_duplicates()
            updated_samp_df_with_pools = pd.concat([sample_df, new_pooled_rows_df], ignore_index=True)
            return updated_samp_df_with_pools
        else:
            return sample_df
    
    def update_rel_cont_id_with_pool_samp(self, sample_df: pd.DataFrame) -> pd.DataFrame:
        # if a subsample of a pooled sample exists in the rel_cont_id column, add the pooled sample to rel_cont_id
        for i, r in sample_df.iterrows():
            rel_cont_ids = str(r[self.faire_rel_cont_id_col_name]).split(' | ')
            updated_rel_ids = []
            for rel_id in rel_cont_ids:
                # check if rel id exists as a subsample of a pooled sample
                for pooled_info in self.pooled_samps_dict:
                    if rel_id in pooled_info['samps_that_were_pooled']:
                        pooled_samp_name = pooled_info['pooled_samp_name']
                        updated_rel_ids.append(pooled_samp_name)
                    else:
                        updated_rel_ids.append(rel_id)
            # remove duplicated rel_ids that happend in the loop with set 
            updated_rel_ids = ' | '.join(list(set(updated_rel_ids)))
            sample_df.at[i, self.faire_rel_cont_id_col_name] = updated_rel_ids
        return sample_df

    def _filter_samp_df_by_samp_name(self, samp_df: pd.DataFrame, associated_seq_df: pd.DataFrame) -> pd.DataFrame:
        # filter sample dataframe to keep only rows where sample names exist in associated seq data frame
        
        missing_samples = []
        # Get unique sample name values from associated Dataframe
        exp_valid_samps = set(associated_seq_df[self.faire_sample_name_col].unique())

        # Also valid samples are pooled subsamples that won't necessarily exist in the experiment run metadata
        other_valid_samps = set()
        for pooled_info in self.pooled_samps_dict:
            for samp in pooled_info['samps_that_were_pooled']:
                if samp in set(samp_df[self.faire_sample_name_col]):
                    other_valid_samps.add(samp)

        # combine valid samps
        valid_samp_names = exp_valid_samps | other_valid_samps

        # Identify missing sample names (those in primary but not in associated seq runs)
        cruise_sample_samp_names = samp_df['samp_name'].unique()
        missing_samples = [name for name in cruise_sample_samp_names if name not in valid_samp_names]
        
        filtered_df = samp_df[samp_df['samp_name'].isin(valid_samp_names)].copy()

        return filtered_df, missing_samples
    
    def create_positive_sample_map(self, run_df, sample_df):
        # Create a dictionary of positive samps: associated samps
        positive_sample_map = {}
        
        grouped = run_df.groupby('assay_name')
        
        for assay_name, group in grouped:
            # Find samples with positive in their name
            positive_samples = group[group[self.faire_sample_name_col].str.contains('POSITIVE', case=False)][self.faire_sample_name_col].unique()
            # Get all sample names in this gropu that are also in the sample_df
            valid_samples = group[group[self.faire_sample_name_col].isin(sample_df[self.faire_sample_name_col])][self.faire_sample_name_col].unique().tolist()

            # for each positive sample add an entry to dictionary
            for pos_sample in positive_samples:
                other_samples = [sample for sample in valid_samples if sample != pos_sample]
                if other_samples:
                    positive_sample_map[pos_sample] = other_samples

        return positive_sample_map
    
    def add_pos_samps_to_rel_cont_id(self, metadata_row:pd.Series, positive_sample_map: dict) -> str:
        # Add the positive sample names to the rel_cont_id and returns the string list with | separator
        sample_name = metadata_row[self.faire_sample_name_col]
        
        # to account for na values - I think these are just empty rows
        try:
            rel_cont_ids = metadata_row[self.faire_rel_cont_id_col_name].split(' | ')
        except:
            rel_cont_ids = []

        for pos_samp, associated_samps in positive_sample_map.items():
            if sample_name in associated_samps:
                rel_cont_ids.append(pos_samp)

        # if 'not applicable' in rel_cont_id and len > 0 then remove that str.
        if len(rel_cont_ids) > 1:
            filtered_list = [id for id in rel_cont_ids if "not applicable" not in id.lower()]
        else:
            filtered_list = rel_cont_ids

        # Remove any nans from list - this is true for pooled samples that were added and no rel_cont_id existed
        filtered_list.remove('nan') if 'nan' in filtered_list else filtered_list
        # rejoin into format
        rel_cont_ids = ' | '.join(filtered_list)
        
        return rel_cont_ids

    def get_pos_cont_type(self, pos_cont_sample_name: str) -> str:

        if 'ferret' in pos_cont_sample_name.lower():
            pos_cont_type = 'PCR positive of Ferret genomic DNA'
        elif 'camel' in pos_cont_sample_name.lower():
            pos_cont_type = 'PCR positive of Camel genomic DNA'
        else:
            # Get the marker from the positive control sample name (e.g. run2.ITS1.POSITIVE - this will give you ITS1)
            shorthand_marker = pos_cont_sample_name.split('.')[1]

            g_block = marker_shorthand_to_pos_cont_gblcok_name.get(
                shorthand_marker)

            pos_cont_type = f"PCR positive of synthetic DNA. Gblock name: {g_block}"

        return pos_cont_type
   
    def add_positive_samps_to_samp_df(self, samp_df: pd.DataFrame, positive_sample_map: dict) -> pd.DataFrame:
        # Add positive samples as their own rows in the data frame
        
        new_rows = []
        for pos_samp, samples in positive_sample_map.items():
            # check that at least one sample matches from pos_sample_map in sample_df and if it does, add positive sample to sample df
            matching_samps = samp_df[samp_df[self.faire_sample_name_col].isin(samples)]
            if not matching_samps.empty:
                pos_cont_type = self.get_pos_cont_type(pos_cont_sample_name=pos_samp)
                new_pos_samp_row = {
                    self.faire_sample_name_col: pos_samp, 
                    self.faire_sample_category_col_name: 'positive control', 
                    self.faire_pos_cont_type_col: pos_cont_type
                    }
                new_rows.append(new_pos_samp_row)

        if new_rows:
            new_pos_rows_df = pd.DataFrame(new_rows)
            # drop any duplicate positive control rows
            new_df = new_pos_rows_df.drop_duplicates()
            # add positive samples to samp df
            updated_samp_df_with_pos = pd.concat([samp_df, new_df], ignore_index=True)

        return updated_samp_df_with_pos

    def add_pcr_rows_to_samp_df(self, sample_df: pd.DataFrame, exp_run_df: pd.DataFrame) -> pd.DataFrame:
    # Add PCR rows to sample df and duplicates data
        
        # Extract the sample base name
        for i, r in exp_run_df.iterrows():
            sample_name = r[self.faire_sample_name_col]
            if '.PCR' in sample_name:
                exp_run_df.at[i, 'base_samp_name'] = sample_name.split('.PCR')[0]

        # Find which rows in sample_df need to be expanded to be duplicated because of PCR replicates
        transform_dict = {}
        if 'base_samp_name' in exp_run_df.columns:
            for base_name in sample_df[self.faire_sample_name_col]:
                matches = exp_run_df[exp_run_df['base_samp_name'] == base_name][self.faire_sample_name_col].tolist()
                if matches:
                    transform_dict[base_name] = set(matches)
       
            # Add to the sample_df data frame
            transformed_rows = []
            for i, r in sample_df.iterrows():
                base_name = r[self.faire_sample_name_col]
                if base_name in transform_dict:
                    for extended_name in transform_dict[base_name]:
                        new_row = r.copy()
                        new_row[self.faire_sample_name_col] = extended_name
                        transformed_rows.append(new_row)
                else:
                    transformed_rows.append(r)

            samp_df_updated_for_pcr = pd.DataFrame(transformed_rows)
            
            return samp_df_updated_for_pcr
        else:
            return sample_df
    
    def add_assay_name_to_samp_df(self, samp_df: pd.DataFrame, exp_run_df: pd.DataFrame) -> pd.DataFrame:
        # Gets a list of assays related to each sample and adds to assay_name in samp_df

        samp_assay_dict = exp_run_df.groupby(self.faire_sample_name_col)[self.faire_assay_name_col].agg(lambda x: ' | '.join(list(x.unique()))).to_dict()

        # Account for pooled samps - add the same assays as their pooled parent sample
        for pooled_info in self.pooled_samps_dict:
            if pooled_info['pooled_samp_name'] in samp_assay_dict:
                assays = samp_assay_dict.get(pooled_info['pooled_samp_name'])
                for samp in pooled_info['samps_that_were_pooled']:
                    samp_assay_dict[samp] = assays

        samp_df[self.faire_assay_name_col] = samp_df[self.faire_sample_name_col].map(samp_assay_dict)

        return samp_df

    def fill_empty_vals(self, df: pd.DataFrame):
        modified_df = df.copy()

        control_mask = df[self.faire_sample_category_col_name].str.contains('negative|positive', case=False, na=False)
        sample_mask = df[self.faire_sample_category_col_name].str.contains('sample', case=False, na=False)

        for col in df.columns:
            empty_conditions = df[col].isna() | (df[col] == '') | (df[col].astype(str) == 'nan')

            # Fill empty control samples
            control_empty = control_mask & empty_conditions
            modified_df.loc[control_empty, col] = 'not applicable: control sample'

            # fill empty sample values
            sample_empty = sample_mask & empty_conditions
            modified_df.loc[sample_empty, col] = 'missing: not collected'

        return modified_df
    
    def filter_exp_run_metadata(self, final_sample_df: pd.DataFrame, exp_run_df: pd.DataFrame) -> pd.DataFrame:
        # Filters the experiment run metadata df to remove rows with samples that don't exist in the sample Metadata
        # This should be done at the very end of processing sample and experiment run metadata because the sample metadata first needs to be filtered
        # by comparing to the experiment run metadata. This step removes rows of samples from other projects (E.g. RC0083 of EcoFoci) that has a run shared with EcoFoci cruises (run1 ,2 and 3)

        
        # Get list of reference names
        reference_names = set(final_sample_df[self.faire_sample_name_col].unique())

        # Create mask of rows to keep
        mask = exp_run_df[self.faire_sample_name_col].isin(reference_names)

        # filter exp_run_df
        exp_run_df_filtered = exp_run_df[mask]

        # Get dropped samples
        dropped_samples = exp_run_df.loc[~mask, self.faire_sample_name_col].unique()
        
        return exp_run_df_filtered, dropped_samples
    
    def retrive_github_bebop(self, owner: str, repo: str, file_path: str):
        url = f"https://api.github.com/repos/{owner}/{repo}/contents/{file_path}"

        headers = {
            'Authorization': self.gh_token,
            'Accept': 'application/vnd.github.v3+json'
        }

        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            data = response.json()

            if 'content' in data:
                # Decode base64 to get the raw markdown file
                base64_content = data['content'].replace('\n', '').replace(' ', '')
                markdown_content = base64.b64decode(base64_content).decode('utf-8')
                
                with tempfile.NamedTemporaryFile(mode='w', suffix='.md', delete=True, encoding='utf-8') as temp_file:
                    temp_file.write(markdown_content)
                    temp_file_path = temp_file.name
                    post = self.load_beBop_yaml_terms(path_to_bebop=temp_file_path)
                    return post.metadata
        
        except requests.exceptions.RequestException as e:
            print(f"Error fetching bebop: {e}")
            return None
    
    def load_assay_level_metadata_to_excel(self, final_exp_run_df: pd.DataFrame) -> None:

        # create list of assays from experiment_run_metadata so only grabbing assays that are actually in the project
        assays_in_proj = final_exp_run_df[self.faire_assay_name_col].unique().tolist()

        col_index = 0
        for assay, bebops in project_pcr_library_prep_mapping_dict.items():
            if assay in assays_in_proj:
                # Get pcr be bop_dict
                pcr_owner = bebops['pcr_bebop']['owner']
                pcr_repo = bebops['pcr_bebop']['repo']
                pcr_file_path = bebops['pcr_bebop']['file_path']
                pcr_bebop = self.retrive_github_bebop(owner=pcr_owner, repo=pcr_repo, file_path=pcr_file_path)
            
                # Get library preparation bebop dict
                lib_owner = bebops['library_bebop']['owner']
                lib_repo = bebops['library_bebop']['repo']
                lib_file_path = bebops['library_bebop']['file_path']
                lib_bebop = self.retrive_github_bebop(owner=lib_owner, repo=lib_repo, file_path=lib_file_path)
                
                assay_col_num = self.project_sheet_assay_start_col_num + col_index
                self.map_pcr_library_prep_to_excel(pcr_bebop, lib_bebop, assay_col_num)
                col_index = col_index + 1
                print(f"Saved assay {assay} to projectMetadata!")
            else:
                print(f"assay {assay} is not in this project - skipping adding it to ProjectMetadata")
            
    def map_pcr_library_prep_to_excel(self, pcr_bebop_dict: dict, library_prep_bebop_dict: dict, assay_col_num: int) -> None:
        # maps assay and library prep to projectMetadata sheet and saves in excel
        workbook = openpyxl.load_workbook(self.final_faire_template_path)
        worksheet = workbook[self.faire_project_metadata_sheet_name]

        # Get the column numbers
        max_row = worksheet.max_row

        # Map dictionaries to excel
        for row in range(2, max_row + 1):
            term_name_cell = worksheet.cell(row=row, column=self.project_sheet_term_name_col_num)
            term_name = term_name_cell.value

            if term_name and term_name in pcr_bebop_dict:
                pcr_cell = worksheet.cell(row=row, column=assay_col_num)
                pcr_cell.value = pcr_bebop_dict[term_name]

            if term_name and term_name in library_prep_bebop_dict:
                lib_prep_cell = worksheet.cell(row=row, column=assay_col_num)
                lib_prep_cell.value = library_prep_bebop_dict[term_name]

        workbook.save(self.final_faire_template_path)
        workbook.close()
    
    def get_sun_times_from_iso(self, metadata_row: pd.Series):
        # Gets the sunrise and sunset from ISO dattime str, return sunrise, sunset
        lat = metadata_row[self.faire_latitude_col]
        lon = metadata_row[self.faire_longitude_col]
        event_date = metadata_row[self.faire_eventDate_col]

        sample_cat = metadata_row[self.faire_sample_category_col_name]
        
        try:
            if sample_cat == 'sample':
                # Handle both "Z" and "+00:00" UTC formats
                if event_date.endswith('Z'):
                    event_date = event_date[:-1] + '+00:00'

                dt_utc = datetime.fromisoformat(event_date)
                date_obj = dt_utc.date()

                # create location
                location = LocationInfo(latitude=lat, longitude=lon)

            # Calculate sun times in UTC
                s = sun(location.observer, date=date_obj, tzinfo=pytz.UTC)

                # Convert to ISO format with 'Z' suffix
                sunrise_iso = s['sunrise'].isoformat().replace('+00:00', 'Z')
                sunset_iso = s['sunset'].isoformat().replace('+00:00', 'Z')

                sunset_sunrise_method = "sunset and sunrise times calculated using eventDate and python's astral sun library."
            

                return sunrise_iso, sunset_iso, sunset_sunrise_method
            else:
                return 'not applicable: control sample', 'not applicable: control sample', 'not applicable: control sample'
        except ValueError as e:
            print(f"\033[93m{metadata_row[self.faire_sample_name_col]}: {e}\033[0m")
            return 'not applicable'

    def drop_samps_from_faire_rel_column_that_dropped(self, df: pd.DataFrame) -> pd.DataFrame:
        # Drop the sample names from rel_cont_id, or biological_rep_relation that dont' exist as a samp_name in the sample metadata df
        valid_samples = set(df[self.faire_sample_name_col])

        # Also valid samples are pooled subsamples that won't necessarily exist in the experiment run metadata
        other_valid_samps = set()
        for pooled_info in self.pooled_samps_dict:
            for samp in pooled_info['samps_that_were_pooled']:
                if samp in set(df[self.faire_sample_name_col]):
                    other_valid_samps.add(samp)

        def filter_faire_relation_str(faire_rel_col_str: str) -> str:
            # funciton to filter rel_cont_id values
            if 'not applicable' not in faire_rel_col_str:
                valid_ids = [sample for sample in faire_rel_col_str.split(' | ') if sample in valid_samples]
                if valid_ids:
                    return ' | '.join(valid_ids) 
                else:
                    return 'not applicable'
            else:
                return faire_rel_col_str
            
        df[self.faire_rel_cont_id_col_name] = df[self.faire_rel_cont_id_col_name].apply(filter_faire_relation_str)
        df[self.faire_biological_rep_relation_col] = df[self.faire_biological_rep_relation_col].apply(filter_faire_relation_str)

        return df

    def drop_samp_cols_not_meant_for_submission(self, sample_df: pd.DataFrame) -> pd.DataFrame:
        # Drop columns not meant for submission like stations_within_5m
        if self.faire_stations_in_5km_col in sample_df.columns:
            return sample_df.drop(columns=[self.faire_stations_in_5km_col])
        else:
            return sample_df
        

    def load_project_level_metadata_to_excel_and_save_as_csv(self) -> None:
        # Maps the project level metadata to the projectMetadata excel sheet
        
        project_dict = dict(zip(self.project_info_df['faire_field'], self.project_info_df['value']))
        
        # Add mod_date to be date code is ran
        today = date.today()
        today_str = today.strftime('%Y-%m-%d')
        project_dict['mod_date'] = today_str

        # save as csv
        data_dir = os.path.dirname(self.final_faire_template_path)
        self.save_final_df_as_csv(final_df=pd.DataFrame([project_dict]), sheet_name='Sheet1', header=0, csv_path=f"{data_dir}/projectMetadata_{self.project_id}.csv")

        # save as excel
        workbook = openpyxl.load_workbook(self.final_faire_template_path)
        worksheet = workbook[self.faire_project_metadata_sheet_name]

         # Get the column numbers
        max_row = worksheet.max_row

        # Map dictionary to excel
        for row in range(2, max_row + 1):
            term_name_cell = worksheet.cell(row=row, column=self.project_sheet_term_name_col_num)
            term_name = term_name_cell.value

            if term_name and term_name in project_dict:
                pcr_cell = worksheet.cell(row=row, column=self.project_sheet_project_level_col_num)
                pcr_cell.value = project_dict[term_name]

        workbook.save(self.final_faire_template_path)
        workbook.close()

    def save_list_to_csv(self, the_list: list, file_name: str) -> None:
        # Puts a list into a csv and saves
        df = pd.DataFrame(the_list)
        df.to_csv(self.logging_directory+file_name, index=False, header=False)