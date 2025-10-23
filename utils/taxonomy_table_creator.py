from pathlib import Path
import pandas as pd
import hashlib
import openpyxl
from openpyxl.utils import get_column_letter

# TODO: Columns left to add :
    # scientificNameAuthorship
    # taxonID
    # taxonID_db (either NCBI, Bold, or PR2, or WORMS, or GBIF)
# TODO: Add assay name to sheet name when saving to the FAIRe excel file???
# TODO: THis only works for Revamp - will need to adjust/expand for other taxonomy methods

class TaxonomyTableCreator:
    # FAIRE Taxonomy fields
    SEQ_ID = "seq_id"
    DNA_SEQUENCE = "dna_sequence"
    KINGDOM = "kingdom"
    PHYLUM = "phylum"
    CLASS = "class"
    ORDER = "order"
    FAMILY = "family"
    GENUS = "genus"
    SPECIES = "species"
    SPECIFIC_EPITHET = "specificEpithet"
    SCIENTIFIC_NAME = "scientificName"
    SCIENTIFIC_NAME_AUTHORSHIP = "scientificNameAuthorship"
    TAXON_RANK = "taxonRank"
    TAXON_ID_DB = "taxonID_db"
    ACCESSION_ID = "accession_id"
    ACCESSION_ID_REF_DB = "accession_id_ref_db"
    VERBATIM_IDENTFICATION = "verbatimIdentification"
    PERCENT_MATCH = "percent_match"
    PERCENT_QUERY_COVER = "percent_query_cover"
    CONFIDENCE_SCORE = "confidence_score"

    TAXONOMY_COLS = [KINGDOM, PHYLUM, CLASS, ORDER, FAMILY, GENUS, SPECIES]
    
    # data
    TAXONOMY_UNKNOWNS = ["Unknown"] # TODO - REvamp uses Unknown, not sure if there will be other ways of saying this for other dbs.
    
    # REVAMP specific (maybe?)
    REVAMP_ACCESSION_MATCH_COL = "accession"
    REVAMP_PERCENT_MATCH_COL = "percent"
    REVAMP_MATCH_LENGTH_COL = "length"
    REVAMP_ACCESSION_ID_REF_DB = "NCBI Nucleotide"

    # Other
    TAXA_FAIRE_TEMPLATE_PATH = "/home/poseidon/zalmanek/FAIRe-Mapping/taxonomy/taxa_faire_template.xlsx"

    def __init__(self, taxon_tble_file: str, asv_match_db_results_file: str, asv_dna_seq_file: str, final_faire_template_path: str):
        """
        taxon_table_file: The path to the taxonomy table.txt file
        asv_match_db_results_file: The path to the accessions database results. E.g. for revamp its the blast results.
        asv_dna_seq_file: The path to the file with the asv/dna sequences
        """

        self.taxon_table_path = Path(taxon_tble_file)
        self.asv_match_db_results_file_path = Path(asv_match_db_results_file)
        self.asv_dna_seq_file = Path(asv_dna_seq_file)
        self.final_faire_template_path = final_faire_template_path
        self.hash_dna_seq_dict = self._create_asv_hash_dict()
        self.asv_match_db_results_df = self._get_asv_match_db_df()
        self.seq_accession_ids_dict = self._create_accession_ids_dict()
        self.seq_percent_match_dict = self._create_percent_match_dict()
        self.seq_query_cover_dict = self._create_percent_query_cov_dict()
        self.taxon_df = self.get_faire_df()

        self.add_final_df_to_FAIRe_excel(final_faire_df=self.taxon_df)
    
    def get_faire_df(self) -> pd.DataFrame:
        """
        Load the taxon_tble_file as a data frame
        """
        # Load taxonomy text file as a data frame.
        df = self._load_tab_text_file_as_df(file_path=self.taxon_table_path)

        # Rename columns to match FAIRE
        df_cols_renamed = self._rename_kpcofgs_columns(df=df)

        # Replace ASVs with hashes
        df_with_hashes = self._switch_asv_for_hashes(df=df_cols_renamed)

        # First get VerbatimIdentification (because the Species column will be dropped and replaced 
        # with scientific epithet which changes the value)
        df_with_hashes[self.VERBATIM_IDENTFICATION] = df_with_hashes.apply(lambda row: self._get_verabtim_identification(row=row), 
                                                                           axis=1)

        # More Species column to be Scientific Epithet. Drop the species column after
        df_with_hashes[self.SPECIFIC_EPITHET] = df_with_hashes[self.SPECIES].apply(self._get_specific_epithet)

        # Get scientific name
        df_with_hashes[self.SCIENTIFIC_NAME] = df_with_hashes.apply(self._get_scientific_name, axis=1)
        
        # Get taxon_rank
        df_with_hashes[self.TAXON_RANK] = df_with_hashes.apply(self._get_taxon_rank, axis=1)

        # Get accession ids and accession_id_ref_db
        df_with_hashes[self.ACCESSION_ID] = df_with_hashes[self.SEQ_ID].map(self.seq_accession_ids_dict).fillna("not applicable")
        df_with_hashes[self.ACCESSION_ID_REF_DB] = self._get_accession_ref_db()

        # Get percent_match
        df_with_hashes[self.PERCENT_MATCH] = df_with_hashes[self.SEQ_ID].map(self.seq_percent_match_dict).fillna("not applicable")

        # Get percent_query_cover
        df_with_hashes[self.PERCENT_QUERY_COVER] = df_with_hashes[self.SEQ_ID].map(self.seq_query_cover_dict).fillna("not applicable")

        # Get confidence score
        df_with_hashes[self.CONFIDENCE_SCORE] = self._get_confidence_score()
        return df_with_hashes
    
    def _get_asv_match_db_df(self) -> pd.DataFrame:
        """
        Puts the asv_match_db_results_file into a data frame. Replaces 
        ASV strings with md5 hashes.
        """
        df = self._load_tab_text_file_as_df(file_path=self.asv_match_db_results_file_path)

        df.rename(columns={'ASV': self.SEQ_ID}, inplace=True)

        # Replace ASVs with hashes
        df_with_hashes = self._switch_asv_for_hashes(df=df)

        return df_with_hashes
    
    def _load_tab_text_file_as_df(self, file_path: Path) -> pd.DataFrame:
        """
        Loads a text file that is tab delimated as a pandas data frame.
        """
        df = pd.read_csv(file_path, sep='\t')

        return df
    
    def _create_asv_hash_dict(self) -> dict:
        # creates a dictionary like {'ASV1': {'seq': ACGT, 'hash': 'dafadsf'}}
        # replace this in experiment_run_mapper if end up using the same dictionary.
        asv_hash_dict = {}
        current_asv = None
        seq_lines = []

        print(f"Creating ASV hash dict for {self.asv_dna_seq_file}")
        with open(self.asv_dna_seq_file, 'r') as f:
            for line in f:
                line = line.strip()

                if not line: # Skip empty lines
                    continue
                
                # Get the first ASV
                if current_asv == None:
                    if line.startswith('>'):
                        current_asv = line.replace('>', '')
                        seq_lines = []
                    else:
                        seq_lines.append(line)

                # For other ASVs that aren't first row
                else:
                    if line.startswith('>'):
                        # Add the old current asv first
                        seq = "".join(seq_lines)
                        seq_hash = hashlib.md5(seq.encode()).hexdigest()
                        asv_hash_dict[current_asv] = {
                            'seq': seq, 
                            'hash': seq_hash
                            }

                        # Then create new current_asv and reset seq lines
                        current_asv = line.replace('>', '')
                        seq_lines = []
                    else:
                        seq_lines.append(line)
        
        # *** CRITICAL FIX: Process the last ASV after the loop finishes ***
        if current_asv is not None and seq_lines:
            seq = "".join(seq_lines)
            seq_hash = hashlib.md5(seq.encode()).hexdigest()
            asv_hash_dict[current_asv] = {
                'seq': seq, 
                'hash': seq_hash
                }

        return asv_hash_dict

    def _rename_kpcofgs_columns(self, df: pd.DataFrame) -> pd.DataFrame:   
        """
        Renames the columns from the original loaded taxonomy file to match
        FAIRe's colums for Kingdom, Phylum, Class, Order, Family, Genus, Species
        """
        col_renames = {}
        for col in df.columns:
            col = col.strip()
            if col.lower() == 'asv':
                col_renames[col] = self.SEQ_ID
            elif col.lower() == self.KINGDOM:
                col_renames[col] = self.KINGDOM
            elif col.lower() == self.PHYLUM:
                col_renames[col] = self.PHYLUM
            elif col.lower() == self.CLASS:
                col_renames[col] = self.CLASS
            elif col.lower() == self.ORDER:
                col_renames[col] = self.ORDER
            elif col.lower() == self.FAMILY:
                col_renames[col] = self.FAMILY
            elif col.lower() == self.GENUS:
                col_renames[col] = self.GENUS
            elif col.lower() == self.SPECIES:
                col_renames[col] = self.SPECIES

        df.rename(columns=col_renames, inplace=True)
        return df

    def _switch_asv_for_hashes(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Switches the ASV names in column one of the taxonomy df for the 
        corresponding hashes
        """
        # Create flat mapping dictionary of {'ASV_1': 'kasdhfkajdsf'}
        hash_map = {
            seq_id: data['hash'] for seq_id, data in self.hash_dna_seq_dict.items()
        }
        # Replace ASV strings with hashes
        df[self.SEQ_ID] = df[self.SEQ_ID].map(hash_map)
        return df

    def _get_verabtim_identification(self, row: pd.Series) -> str:
        """
        Get the verbatim identification by string together values in 
        taxonomic columns.
        """
        kingdom = row[self.KINGDOM]
        phylum = row[self.PHYLUM]
        tax_class = row[self.CLASS]
        order = row[self.ORDER]
        family = row[self.FAMILY]
        genus = row[self.GENUS]
        species = row[self.SPECIES]

        verbatim_list = [kingdom, phylum, tax_class, order, family, genus, species]
        # Drop na values (but keep any strings the say unkown, or what not)
        verbatim_list = [val for val in verbatim_list if pd.notna(val)]
        verbatim_str = ', '.join(verbatim_list)
        return verbatim_str

    def _get_taxon_rank(self, row: pd.Series) -> str:
        """
        Gets the most specific column name for the most specific
        name in the Scientific name.
        """
        # Iterate through the taxonomic cols in reverse order (most specific first)
        for col in reversed(self.TAXONOMY_COLS):
            if pd.notna(row[col]) and row[col] not in self.TAXONOMY_UNKNOWNS:
                if col == self.SPECIFIC_EPITHET:
                    return "species"
                else:
                    return col
   
        return "missing: not provided"
    
    def _get_scientific_name(self, row: pd.Series) -> str:
        """
        Gets the most scientific name for the most specific 
        value in the taxonomies.
        """
        for col in reversed(self.TAXONOMY_COLS):
            if pd.notna(row[col]) and row[col] not in self.TAXONOMY_UNKNOWNS:
                return row[col]
        
        return "not applicable"

    def _get_specific_epithet(self, species_value: str) -> str:
        """
        The specific epithet column is currently just mapped to the species
        column in the origina taxonomy.txt file, need to edit to fit 
        the speicicEpithet which is the lower case value of the species, 
        everything else that is NA will become 'not applicable'
        """
        
        if pd.isna(species_value) or species_value in self.TAXONOMY_UNKNOWNS:
            return "not applicable"
        else:
            species_words = str(species_value).split()
            if len(species_words) ==2:
                return species_words[1]
            elif len(species_words) > 2 and ('sp_' in species_words or 'cf_' in species_words or 'aff_' in species_words):
                return f"Unclassified {' '. join(species_words[1:])}"
            elif len(species_words) == 3: # Has subspecies
                return species_words[1]
            elif "endosymbiont":
                return "endosymbiont"
            else:
                print(f"\033[35mThere appear to be a weird species structure with value {species_value}.\033[0m")
                return "not applicable" 
                    
    def _create_accession_ids_dict(self) -> dict:
        """
        Creates a dictionary of seq: accession ids
        """    
        accession_dict = self.asv_match_db_results_df.set_index(self.SEQ_ID)[self.REVAMP_ACCESSION_MATCH_COL].to_dict()
        # update list of accession to use |
        for accession in accession_dict.values():
            accession.replace(',', ' |')

        return accession_dict      

    def _get_accession_ref_db(self) -> str:
        """
        Get the accession ref db based on the file names
        """
        if "revamp" in str(self.taxon_table_path).lower():
            accession_ref_db = self.REVAMP_ACCESSION_ID_REF_DB
        else:
            raise ValueError("Have not added functionality for your reference database!")

        return accession_ref_db

    def _create_percent_match_dict(self) -> dict:
        """
        Creates the seq_id: percent_match dictionary
        """
        percent_match_dict = self.asv_match_db_results_df.set_index(self.SEQ_ID)[self.REVAMP_PERCENT_MATCH_COL].to_dict()

        return percent_match_dict       

    def _create_percent_query_cov_dict(self) -> dict:
        """
        Creates a dictionary of seq_id: percent_query_coverage by creating
        a dictionary of the seq_id: match_length. And then finding the 
        length of the DNA sequences
        """
        length_match_dict = self.asv_match_db_results_df.set_index(self.SEQ_ID)[self.REVAMP_MATCH_LENGTH_COL].to_dict()

        # Create dictionary of {seq_id: seq_length}
        seq_length_map = {
            data['hash']: len(data['seq']) for data in self.hash_dna_seq_dict.values()
        }

        percent_query_cov_dict = {}
        for seq_hash, length in length_match_dict.items():
            # Turn lengths from comma separated string into list of integers
            match_lengths = length.split(',')
            match_lengths = [int(s) for s in match_lengths]

            # Divide match lengths by sequence lengths (if more than one value only get the first and last value of the list and give a range)
            percent_query_cov_list = []
            for match_length in match_lengths:
                seq_length = seq_length_map.get(seq_hash, None)
                percent_cov = (match_length/seq_length)*100
                percent_query_cov_list.append(percent_cov)
            percent_query_cov_list = sorted(percent_query_cov_list)
            if len(percent_query_cov_list) > 1:
                percent_query_cov_str = f"{round(percent_query_cov_list[0], 1)} - {round(percent_query_cov_list[-1], 1)}"
            elif len(percent_query_cov_list) == 1:
                percent_query_cov_str = str(round(percent_query_cov_list[0], 1))
            percent_query_cov_dict[seq_hash] = percent_query_cov_str

        return percent_query_cov_dict
                    
    def _get_confidence_score(self):
        """
        Get the confidence score, for REVAMP, just NA
        """   
        if "revamp" in str(self.taxon_table_path).lower():
            return "not applicable"  
        else: 
            raise ValueError("Do not know how to get confidence score for this method! Add functionlity!")

    def add_final_df_to_FAIRe_excel(self, final_faire_df: pd.DataFrame, sheet_name: str = "taxaFinal"):
        # Step 1 load the workbook to preserve formatting
        workbook = openpyxl.load_workbook(self.TAXA_FAIRE_TEMPLATE_PATH)
        sheet = workbook[sheet_name]

        # Step 2: Store original row 2 headers and original row 3 columns
        # This is done before any modifications to identify what was originally in the template.
        original_row2_headers = {}
        original_columns_in_sheet = []
        for col_idx in range(1, sheet.max_column + 1):
            col_name = sheet.cell(row=3, column=col_idx).value
            if col_name:
                original_columns_in_sheet.append(col_name)
                # Store the row 2 value for this column name
                original_row2_headers[col_name] = sheet.cell(row=2, column=col_idx).value
        
        # Step 4: Clear existing headers (rows 2 and 3) and all data from row 4 onwards
        # This ensures a clean slate before writing new, reordered headers and data.
        for row in range(1, sheet.max_row + 1): # Clear starting from row 1 to be safe
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None

        # Step 5: Write the new headers (row 2 and row 3) based on the reordered DataFrame
        for col_idx, col_name in enumerate(final_faire_df.columns, 1):
            col_letter = get_column_letter(col_idx)
            
            # Write column name to row 3 (the main header row)
            sheet[f'{col_letter}3'] = col_name

            # Determine and write row 2 header
            if col_name in original_columns_in_sheet:
                # If the column existed in the original template, use its original row 2 header
                sheet[f'{col_letter}2'] = original_row2_headers.get(col_name)
            else:
                # If it's a new column, set row 2 to "User defined"
                sheet[f'{col_letter}2'] = 'User defined'

        # Step 6: Write the data frame data to the sheet (starting at row 4)
        # Iterate through the reordered DataFrame and write its values to the Excel sheet.
        for row_idx, row_data in enumerate(final_faire_df.values, 4):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx).value = value

        # Step 7: Save the workbook
        workbook.save(self.final_faire_template_path)
        print(f"sheet {sheet_name} saved to {self.final_faire_template_path}!")
                    
                    
                   