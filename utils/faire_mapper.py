from pathlib import Path
import openpyxl
import csv
import frontmatter
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd
# import difflib
import warnings
import yaml
import re
import numpy as np
from .custom_exception import ControlledVocabDoesNotExistError
from .lists import faire_int_cols
import gspread #library that makes it easy for us to interact with the sheet
from google.oauth2.service_account import Credentials
# import requests
# from bs4 import BeautifulSoup

# TODO: Fix sample name - either map from Sample column or add PCR back into Sample_Name column, and a technical replicates column
# TODO: double check neg_cont_type logic after Sean and Zack add these into the run metadata?
# TODO: add pos_cont_type logic after Sean and Zack and positive control to run metadata?
# TODO: check controls and make sure the metadata make sense for those (e.g. samp_collect_device - would this be a CTD Niskin?)
# field_negatives are sterile RO water pre-filled and sealed prior to sample - so a lot of this metadata does not make sense for it.
# TODO: Which other columns should we include that were not included: e.g. Cast_No, Rosette_position, Station, Cruise_ID_short, Cruise_ID_long, sites, technical_reps
# TODO: Split into on sheet per cruise - or keep as one file because in Luke's template he has added assay columns in the Project Metadata (to account for different assays across cruises)
# TODO: add logic for env_local_scale based on depth
# TODO: maybe try to figure out logic for mapping two columns to one (e.g. silicate stuff) in a better way. But will this ever happen again? Happening cause of combining cruise data
# TODO: Fix GeoLoc to pull in location in R scripts (I think I saw Bering/Chulchi Sea somehwere in that data) and add Arctic Ocean in main.
# TODO: add blanks to metadata sheet (not sure if this can be pulled in somewhere besides the run1_4_revamp_metadata)
# TODO: transorm Extraction Date to meet standard
# TODO Someday: Get faire_missing_values from parsing webpage using BeautifulSoup (to allow for changes: https://fair-edna.github.io/guidelines.html#missing-values)
class OmeFaireMapper:
    
    def __init__(self, config_yaml: str):

        self.config_file = self.load_config(config_yaml)
        self.google_sheet_json_cred = self.config_file['json_creds']

        self.mapping_file_FAIRe_column = 'faire_field'
        self.mapping_file_metadata_column = 'source_name_or_constant'
        self.mapping_file_mapped_type_column = 'mapping'
        
        self.faire_template_file=self.config_file['faire_template_file']
        self.drop_down_value_df = self.load_faire_template_as_df(self.faire_template_file, sheet_name='Drop-down values', header=0)
        # self.faire_unit_column_dict = {term: term.replace('_unit', '') for term in self.mapping_df[self.mapping_file_FAIRe_column].values if '_unit' in term}
        self.final_faire_template_path = self.config_file['final_faire_template_path']
        self.faire_sheet_header = 2
        self.exact_mapping = 'exact'
        self.related_mapping = 'related'
        self.constant_mapping = 'constant'
        self.faire_missing_values = ["not applicable: control sample",
                                     "not applicable: sample group",
                                     "not applicable",
                                     "missing: not collected: synthetic construct",
                                     "missing: not collected: lab stock",
                                     "missing: not collected: third party data",
                                     "missing: not collected",
                                     "missing: not provided",
                                     "missing: restricted access: endangered species", 
                                     "missing: restricted access: human-identifiable", 
                                     "missing: restricted access"
                                     ]

    def load_config(self, config_path):
        # Load configuration yaml file
    
        with open(config_path, 'r') as f:
            return yaml.safe_load(f)
            
    def load_faire_template_as_df(self, file_path: Path, sheet_name: str, header: int) -> pd.DataFrame:
        # Load FAIRe excel template as a data frame based on the specified template sheet name
        
        return pd.read_excel(file_path, sheet_name=sheet_name, header=header)
    
    def load_csv_as_df(self, file_path: Path, header=0, sep=',') -> pd. DataFrame:
        # Load csv files as a data frame

        return pd.read_csv(file_path, header=header, sep=sep)
    
    def load_google_sheet_as_df(self, google_sheet_id: str, sheet_name: str, header: int) -> pd.DataFrame:

        scopes = ['https://www.googleapis.com/auth/spreadsheets.readonly']

        creds = Credentials.from_service_account_file(self.google_sheet_json_cred, scopes=scopes)
        client = gspread.authorize(creds)

        sheet = client.open_by_key(google_sheet_id)
        worksheet = sheet.worksheet(sheet_name)
        
        # Get all values
        values = worksheet.get_all_values()
        headers = values[header]
        data_rows = values[header+1:]
        df = pd.DataFrame(data_rows, columns=headers)
        
        return df
    
    def load_beBop_yaml_terms(self, path_to_bebop: str):
        # read BeBOP yaml terms
        with open(path_to_bebop, 'r', encoding='utf-8') as f:
            post = frontmatter.load(f)
            return post
        
    def str_replace_for_samps(self, samp_name: pd.Series) -> str:
        # Fixes sample names in the data frame
        # if sample is part of the DY2012 cruise, will replace any str of DY20 with DY2012
        samp_name = str(samp_name)
        if '_' in samp_name and 'pool' not in samp_name: # osu samp names have _ that needs to be . For example, E62_1B_DY20. Pooled samples keep the underscore
            samp_name = samp_name.replace('_', '.')
        if '.DY20' in samp_name:
           samp_name = samp_name.replace('.DY20', '.DY20-12')
        if '(P10 D2)' in samp_name:
            samp_name = samp_name.replace(' (P10 D2)', '') # for E1875.OC0723 (P10 D2) in Run2
        if '.IB.NO20' in samp_name: # for E265.1B.NO20 sample - in metadata was E265.1B.NO20
            return samp_name.replace('.IB.NO20', '.1B.NO20-01')
        if 'E.2139.' in samp_name:
            return samp_name.replace('E.2139.', 'E2139.') # For E239 QiavacTest, had a . between E and number in metadata
        if 'E687' in samp_name:
            return samp_name.replace('E687', 'E687.WCOA21')
        if '.NC' in samp_name: # If an E was put in front of an NC sample (this happends in some of the extractions e.g. the SKQ21 extractions), will remove the E
            samp_name = samp_name.replace('E.', '')
        if '*' in samp_name:
            samp_name = samp_name.replace('*','')
        if '.SKQ2021' in samp_name:
            samp_name = samp_name.replace('.SKQ2021', '.SKQ21-15S')
        if '.NO20' in samp_name:
            samp_name = samp_name.replace('.NO20', '.NO20-01')
        if 'Mid.NC.SKQ21' in samp_name:
            samp_name = samp_name.replace('Mid.NC.SKQ21', 'MID.NC.SKQ21-15S')
        if '.DY2206' in samp_name:
            samp_name = samp_name.replace('.DY2206', '.DY22-06')
        if '.DY2209' in samp_name:
            samp_name =  samp_name.replace('.DY2209', '.DY22-09')
        if '.DY2306' in samp_name:
            samp_name =  samp_name.replace('.DY2306', '.DY23-06')
        if 'E2030.NC' == samp_name:
            samp_name = 'E2030.NC.SKQ23-12S'

        return samp_name
    
    def extract_controlled_vocab(self, faire_attribute: str) -> list:
        
        # filter dataframe by the FAIRe attribute and get all allowable terms
        filtered_row = self.drop_down_value_df[self.drop_down_value_df['term_name'] == faire_attribute]
        # Get a list of columns that have the vocabulary in them
        vocab_columns = [col for col in filtered_row.columns if ('vocab' in col and '_' not in col)]
        # Get a list of the controlled vocabulary by FAIRe attribute term name
        controlled_vocab = [vocab_word for vocab_word in filtered_row[vocab_columns].values.flatten() if pd.notna(vocab_word)]

        return controlled_vocab
     
    def check_cv_word(self, value: str, faire_attribute: str) -> dict:
        # Check a word in a list of the controlled voabulary for a FAIRe attribute to see if it exists (accounts for any updates)
        # and if it does, will add to the new row.
        
        controlled_vocab = self.extract_controlled_vocab(faire_attribute=faire_attribute)
        # make static_value a list (to account for more than one value as the static_value)

        value = value.split(' | ') if '|' in value else [value]

        for word in value:
            if word not in controlled_vocab and word not in self.faire_missing_values and 'other:' not in word:
                warnings.warn(f'The following {faire_attribute} does not exist in the FAIRe standard controlled vocabulary: {word}, the allowed values are {controlled_vocab}')
            else:
                new_value = ' | '.join(value)
                return new_value
    
    def apply_exact_mappings(self, metadata_row, faire_col):

        if faire_col in self.drop_down_value_df['term_name'].values:
            metadata_row = self.check_cv_word(value=metadata_row, faire_attribute=faire_col)
        else:
            return metadata_row

        return metadata_row
    
    def apply_static_mappings(self, faire_col: str, static_value) -> dict:
        # returns static_value for row
        # TODO: need to adjust this for controls that weren't necessarily collected in the wild (e.g. habitat_natural_articicial_0_1, 
        # geo_loc_name?, env scales, etc.)

        # check controlled vocabulary if column uses controlled vocabulary
        if faire_col in self.drop_down_value_df['term_name'].values:
            static_value = self.check_cv_word(value=static_value, faire_attribute=faire_col)
        # elif faire_col == 'geo_loc_name':
        #     self.check_and_add_geo_loc(formatted_geo_loc=static_value, new_row=new_row, faire_col=faire_col)
        return static_value
    
    def map_using_two_or_three_cols_if_one_is_na_use_other(self, metadata_row: pd.Series, desired_col_name: str, 
                                                  use_if_na_col_name: str, transform_use_col_to_date_format=False,
                                                  use_if_second_col_is_na: str = None) -> datetime:
        # If a faire column maps to two columns because there is data missing from the desired column,
        # and will need to use data from another column in that case, then return the other columns data
        # only works if mapping is exact for columns
        # Make date = true if use_if_na_col_name needs to be transformed from 

        if transform_use_col_to_date_format == False:
            # If desired col name value is not na
            if pd.notna(metadata_row[desired_col_name]) and metadata_row[desired_col_name] != '':
                return metadata_row[desired_col_name]
            elif pd.notna(metadata_row[use_if_na_col_name]) and metadata_row[use_if_na_col_name] != '':
                return metadata_row[use_if_na_col_name]
            else:
                try:
                    return metadata_row[use_if_second_col_is_na]
                except:
                    return ''

        else:
            if pd.notna(metadata_row[desired_col_name]) and metadata_row[desired_col_name] != '':
                return self.convert_date_to_iso8601(date=metadata_row[desired_col_name])
            elif pd.notna(metadata_row[use_if_na_col_name]) and metadata_row[use_if_na_col_name] != '':
                return self.convert_date_to_iso8601(date=metadata_row[use_if_na_col_name])
            elif pd.notna(metadata_row[use_if_second_col_is_na]) and metadata_row[use_if_second_col_is_na] != '':
                return self.convert_date_to_iso8601(date=metadata_row[use_if_second_col_is_na])
            
                  
    def convert_date_to_iso8601(self, date: str) -> datetime:
        # converts strings from 2021/11/08 00:00:00 to iso8601 format  to 2021-11-08T00:00:00Z
        # also converts strings from 5/1/2024 to 2024-01-05T00:00:00Z
        # And coverts 2024-05-01 to 2024-01-05T00:00:00Z
        # Also handles years like 0022 and corrects them to 2022
        has_time_component = False

        date = str(date)
        
        if date in  ['None', 'nan', 'missing: not collected', '', 'missing: not provided']:
            return "missing: not provided"

        # 1. Handle full ISO 8601 format (YYYY-MM-DDTHH:MM:SSZ) and return immediately
        try: 
            # Use the correct format including T and Z
            datetime.strptime(date, "%Y-%m-%dT%H:%M:%SZ")
            return date
        except ValueError: 
            pass # Not that format, continue to check others

        # 2. Check for other supported formats
        
        # Format 2021/11/08 00:00:00
        if "/" in date and ":" in date: 
            dt_obj = datetime.strptime(date, "%Y/%m/%d %H:%M:%S")
            has_time_component = True
            
        # Format 5/1/2024 or 5/1/24
        elif "/" in date and ":" not in date: 
            try: # 5/1/2024 format
                dt_obj = datetime.strptime(date, "%m/%d/%Y")
            except ValueError:
                try: # foramt 5/1/24
                    dt_obj = datetime.strptime(date, "%m/%d/%y")
                except ValueError:
                    raise ValueError(f"Unsupported slash-separated dae format: {date}")
        
        # --- FIX APPLIED HERE: Handle all dash-separated formats gracefully ---
        elif "-" in date:
            if ':' in date:
                # 2.1. Try ISO-like T-separated time (e.g., 2023-04-24T08:51:00)
                try:
                    dt_obj = datetime.strptime(date, "%Y-%m-%dT%H:%M:%S")
                    has_time_component = True
                except ValueError:
                    # 2.2. Fallback: Try space-separated time (e.g., 2023-04-24 08:51:00)
                    try:
                        dt_obj = datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
                        has_time_component = True
                    except ValueError:
                        # Failed both time formats, raise error
                        raise ValueError(f"Unsupported dash-separated date/time format: {date}")
            else:
                # 2.3. Date-only format (e.g., 2024-04-10)
                dt_obj = datetime.strptime(date, "%Y-%m-%d")
        # ---------------------------------------------------------------------

        else:
            raise ValueError(f"Unsupported date format: {date}")
        
        # Correct years that are clearly wrong (like 0022 -> 2022)
        if dt_obj.year < 100:
            corrected_year = 2000 + dt_obj.year
            dt_obj = dt_obj.replace(year=corrected_year)

        # Only add time component if it was in the original string
        if has_time_component:
            return dt_obj.strftime("%Y-%m-%dT%H:%M:%SZ")
        else:
            return dt_obj.strftime("%Y-%m-%d")

    def fix_int_cols(self, df:pd.DataFrame) -> pd.DataFrame:
        # converts columns that are int to so will not save as float. May need to update list in .lists
        df = df.copy()
      
        # Add WOCE that may have been missed
        cols_to_convert = []
        cols_to_convert.extend(faire_int_cols)
        cols_to_convert.extend([col for col in df.columns if 'WOCE' in col])
        cols_to_convert = list(set(cols_to_convert))

        for col in cols_to_convert:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: x if pd.isna(x) or x in self.faire_missing_values else int(float(x))) 

        return df
    
    def reorder_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        # orders the final columns so unit and method columns are next to their corresponding fields
        original_cols = df.columns.tolist()
        reordered_cols = []
        processed_cols = set()

        special_cases = {
            'maximumDepthInMeters': ['DepthInMeters_method'],
            'sunset_time_utc': ['sunset_sunrise_method'],
            'niskin_id': ['niskin_WOCE_flag']
        }
        
        for col in original_cols:
            if col in processed_cols:
                continue

            if col.endswith('_unit') or col.endswith('_units') or col.endswith('_method') or col.endswith('_standard_deviation') or col.endswith('_WOCE_flag'):
                continue

            is_special_related_col = False
            for main_col, related_cols in special_cases.items():
                if col in related_cols:
                    is_special_related_col = True
                    break

            if is_special_related_col:
                continue

            # This is a main column, add it first
            reordered_cols.append(col)
            processed_cols.add(col)

            if col in special_cases:
                for special_col in special_cases[col]:
                    if special_col in original_cols:
                        reordered_cols.append(special_col)
                        processed_cols.add(special_col)

            # Look for corresponding unit col (both _unit and _units)
            unit_col = None
            if f"{col}_unit" in original_cols:
                unit_col = f"{col}_unit"
            elif f"{col}_units" in original_cols:
                unit_col = f"{col}_units"

            if unit_col:
                reordered_cols.append(unit_col)
                processed_cols.add(unit_col)

            method_col = f"{col}_method"
            if method_col in original_cols:
                reordered_cols.append(method_col)
                processed_cols.add(method_col)

            standard_dev_col = f"{col}_standard_deviation"
            if standard_dev_col in original_cols:
                reordered_cols.append(standard_dev_col)
                processed_cols.add(standard_dev_col)

            woce_flag_col = f"{col}_WOCE_flag"
            if woce_flag_col in original_cols:
                reordered_cols.append(woce_flag_col)
                processed_cols.add(woce_flag_col)

        for col in original_cols:
            if col not in processed_cols:
                reordered_cols.append(col)

        return df[reordered_cols]

    def save_final_df_as_csv(self, final_df: pd.DataFrame, sheet_name: str, header: int, csv_path: str) -> None:
        
        try:
            faire_template_df = self.load_faire_template_as_df(file_path=self.faire_template_file, sheet_name=sheet_name, header=header)

            faire_final_df = pd.concat([faire_template_df, final_df], ignore_index=True)
        except: # for project metadata (who doesn't have colun names as headers)
            faire_final_df = final_df

        faire_final_int_fixed = self.fix_int_cols(df=faire_final_df)

        faire_final_df_reorderd = self.reorder_columns(df=faire_final_int_fixed)

        faire_final_df_reorderd.to_csv(csv_path, quoting=csv.QUOTE_NONNUMERIC, index=False)

    def add_final_df_to_FAIRe_excel(self, excel_file_to_read_from: str, sheet_name: str, faire_template_df: pd.DataFrame):
        # Step 1 load the workbook to preserve formatting
        workbook = openpyxl.load_workbook(excel_file_to_read_from)
        sheet = workbook[sheet_name]

        # Step 2: Store original row 2 headers and original row 3 columns
        # This is done before any modifications to identify what was originally in the template.
        original_row2_headers = {}
        original_columns_in_sheet = []
        for col_idx in range(1, sheet.max_column + 1):
            col_name = sheet.cell(row=3, column=col_idx).value
            if col_name:
                original_columns_in_sheet.append(col_name)
                # Store the row 2 value for this column name
                original_row2_headers[col_name] = sheet.cell(row=2, column=col_idx).value

        # Step 3: Reorder the input DataFrame's columns and fix the ints
        reordered_faire_template_df = self.reorder_columns(faire_template_df)
        faire_ints_fixed_df = self.fix_int_cols(df=reordered_faire_template_df)
        
        # Step 4: Clear existing headers (rows 2 and 3) and all data from row 4 onwards
        # This ensures a clean slate before writing new, reordered headers and data.
        for row in range(1, sheet.max_row + 1): # Clear starting from row 1 to be safe
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None

        # Step 5: Write the new headers (row 2 and row 3) based on the reordered DataFrame
        for col_idx, col_name in enumerate(faire_ints_fixed_df.columns, 1):
            col_letter = get_column_letter(col_idx)
            
            # Write column name to row 3 (the main header row)
            sheet[f'{col_letter}3'] = col_name

            # Determine and write row 2 header
            if col_name in original_columns_in_sheet:
                # If the column existed in the original template, use its original row 2 header
                sheet[f'{col_letter}2'] = original_row2_headers.get(col_name)
            else:
                # If it's a new column, set row 2 to "User defined"
                sheet[f'{col_letter}2'] = 'User defined'

        # Step 6: Write the data frame data to the sheet (starting at row 4)
        # Iterate through the reordered DataFrame and write its values to the Excel sheet.
        for row_idx, row_data in enumerate(faire_ints_fixed_df.values, 4):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx).value = value

        # Step 7: Save the workbook
        workbook.save(self.final_faire_template_path)
        print(f"sheet {sheet_name} saved to {self.final_faire_template_path}!")
       


        

       
    

    
