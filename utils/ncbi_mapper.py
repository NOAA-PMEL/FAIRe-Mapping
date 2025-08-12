import pandas as pd
import shutil
import frontmatter
import requests
import base64
import tempfile
import re
import numpy as np
from .lists import faire_to_ncbi_units, ncbi_faire_to_ncbi_column_mappings_exact, ncbi_faire_sra_column_mappings_exact
from openpyxl import load_workbook
from pathlib import Path

# TODO: add title to get_sra_df method once hear back from Sean
# TODO: Figure out what is going on when submitting PCR samples - added a custom column to differentiate them and the submitter isn't recognizing it.
# TODO: check on pos_cont_type and neg_cont_type, and sample_Title in ncbi sample mappings stuff?
# TODO: figure out title for SRA mappings - need to ask Sean and Zack
# TODO: switch out manually loading library prep bebop to use the github method (see ProjectMapper method)

class NCBIMapper:

    ncbi_sample_excel_template_path = '/home/poseidon/zalmanek/FAIRe-Mapping/MIMARKS.survey.water.6.0.xlsx'
    ncbi_sra_excel_template_path = '/home/poseidon/zalmanek/FAIRe-Mapping/SRA_metadata.xlsx'
    ncbi_organism = "marine metagenome"
    ncbi_sample_sheet_name = "MIMARKS.survey.water.6.0"
    ncbi_sra_sheet_name = 'SRA_data'
    ncbi_sample_header = 11
    ncbi_sra_header = 0
    ncbi_library_strategy = 'AMPLICON'
    ncbi_library_source = 'METAGENOMIC'
    ncbi_library_selection = 'PCR'
    ncbi_library_layout = 'Paired-end'
    ncbi_file_type = 'fastq'
    ncbi_bioprojec_col_name = 'bioproject_accession'
    faire_missing_values = ["not applicable: control sample",
                            "not applicable: sample group",
                            "not applicable",
                            "missing: not collected: synthetic construct",
                            "missing: not collected: lab stock",
                            "missing: not collected: third party data",
                            "missing: not collected",
                            "missing: not provided",
                            "missing: restricted access: endangered species", 
                            "missing: restricted access: human-identifiable", 
                            "missing: restricted access"
                            ]
    faire_samp_name_col = "samp_name"
    faire_stations_in_5km_col = "station_ids_within_5km_of_lat_lon"
    faire_associated_seqs_col_name = "associatedSequences"
    ncbi_samp_name_col = "*sample_name"

    def __init__(self, final_faire_sample_metadata_df: pd.DataFrame, 
                 final_faire_experiment_run_metadata_df: pd.DataFrame,
                 ncbi_sample_excel_save_path: str,
                 ncbi_sra_excel_save_path: str, 
                 library_prep_dict: dict):
        # Initialize with the final faire_sample_metadata_df and final_faire_experiment_run_metadata_df
        # Can get these by initializing ProjectMapper and running process_sample_run_data method. Be sure to include config file
        # see run4/ncbi_mapping/ncbi_main.py for example

        self.faire_sample_df = self.clean_samp_df(df=final_faire_sample_metadata_df)
        self.faire_experiment_run_df = final_faire_experiment_run_metadata_df
        self.ncbi_sample_template_df = self.load_ncbi_template_as_df(file_path=self.ncbi_sample_excel_template_path, sheet_name=self.ncbi_sample_sheet_name, header=self.ncbi_sample_header)
        self.ncbi_sample_excel_save_path = Path(ncbi_sample_excel_save_path)
        self.ncbi_sra_excel_save_path = Path(ncbi_sra_excel_save_path)
        self.library_prep_bebop = self.retrive_github_bebop(owner=library_prep_dict.get('owner'), repo=library_prep_dict.get('repo'), file_path = library_prep_dict.get('file_path'))
        self.ncbi_bioproject_dict = self.create_ncbi_accession_dict_project_and_biosample(id_prefix='PRJNA')
        self.ncbi_biosample_dict = self.create_ncbi_accession_dict_project_and_biosample(id_prefix='SAMN')
        self.ncbi_srr_dict = self.create_ncbi_accession_dict_project_and_biosample(id_prefix='SRR')

    def create_ncbi_submission(self):
        # Will output two excel files

        # Create NCBI Sample Template and fill out and save
        final_ncbi_sample_df = self.get_ncbi_sample_df()
        self.save_to_excel_template(template_path=self.ncbi_sample_excel_template_path,
                                                    ncbi_excel_save_path=self.ncbi_sample_excel_save_path, 
                                                    sheet_name=self.ncbi_sample_sheet_name,
                                                    header=self.ncbi_sample_header,
                                                    final_ncbi_df=final_ncbi_sample_df)
        
        # Creat NCBI SRA Template and fill out and save
        final_sra_df = self.get_sra_df()
        self.save_to_excel_template(template_path=self.ncbi_sra_excel_template_path,
                                    ncbi_excel_save_path=self.ncbi_sra_excel_save_path,
                                    sheet_name=self.ncbi_sra_sheet_name,
                                    header=self.ncbi_sra_header,
                                    final_ncbi_df=final_sra_df)
    
    def create_osu_ncbi_submission(self):
        
        exp_run_groups, samp_groups = self.split_osu_dfs_by_submission_type()

        for submission_type, sample_df in samp_groups.items():
            # overwrite faire_experiment_run_df and faire_sample_df
            self.faire_experiment_run_df = exp_run_groups.get(submission_type)
            self.faire_sample_df = sample_df


            # Create new excel file save path: 
            new_dir = self.ncbi_sample_excel_save_path.parent / submission_type
            new_sample_file_name = f"{submission_type}_{self.ncbi_sample_excel_save_path.name}"
            new_sample_path = new_dir / new_sample_file_name
            # create directory if it doesn't already exist
            new_dir.mkdir(parents=True, exist_ok=True)

            # process new dfs for ncbi submission
            # First sample df
            final_ncbi_sample_df = self.get_ncbi_sample_df()
            self.save_to_excel_template(template_path=self.ncbi_sample_excel_template_path,
                                                    ncbi_excel_save_path=new_sample_path, 
                                                    sheet_name=self.ncbi_sample_sheet_name,
                                                    header=self.ncbi_sample_header,
                                                    final_ncbi_df=final_ncbi_sample_df)

        for submission_type, exp_run_df in exp_run_groups.items():
            
            new_dir = self.ncbi_sample_excel_save_path.parent / submission_type
            new_expRun_file_name = f"{submission_type}_{self.ncbi_sra_excel_save_path.name}"
            new_expRun_path = new_dir/ new_expRun_file_name
            self.faire_experiment_run_df = exp_run_df
            
            final_sra_df = self.get_sra_df()
            self.save_to_excel_template(template_path=self.ncbi_sra_excel_template_path,
                                    ncbi_excel_save_path=new_expRun_path,
                                    sheet_name=self.ncbi_sra_sheet_name,
                                    header=self.ncbi_sra_header,
                                    final_ncbi_df=final_sra_df)

    def split_osu_dfs_by_submission_type(self):
        # splits the self.samp_df and self.exp_run_df into separate data frames by submission type (e.g. single_direct, nan, etc.)
        # and returns to dictionary, the samp_groups and the exp_run_groups
        # replace NaN values for submission_type with 'NEW'
        
        mapped_biosamp_accessions = self.faire_experiment_run_df[self.faire_samp_name_col].map(self.ncbi_biosample_dict)
        has_biosamp_accession_mask = mapped_biosamp_accessions.notna()
        self.faire_experiment_run_df['group_key'] = np.where(
            has_biosamp_accession_mask,
            'has_biosamp_accession',
            'no_biosamp_accession'
        )
        
        # Step 1: split exp_run_df by submission type (inlcuding NaN or empty)
        exp_run_groups = {}
        for ncbi_biosample_status, group in self.faire_experiment_run_df.groupby('group_key', dropna=False):
            exp_run_groups[ncbi_biosample_status] = group

        # Step 2: Get the unique group/key ncbi_submission statuses to know how to split sample_df
        ncbi_biosample_statuses = self.faire_experiment_run_df['group_key'].unique() 

        # Step 3: split sample_df based on which sample names correspond to each submission value
        samp_groups = {}
        for ncbi_biosample_status in ncbi_biosample_statuses:
            # Get samp names that correspond to this submission value
            samp_names_for_submission = self.faire_experiment_run_df[self.faire_experiment_run_df['group_key'] == ncbi_biosample_status][self.faire_samp_name_col].unique()

            # filter samp_df to only include rows with these samp_names
            samp_groups[ncbi_biosample_status] = self.faire_sample_df[self.faire_sample_df[self.faire_samp_name_col].isin(samp_names_for_submission)]

        return exp_run_groups, samp_groups
    
    def load_ncbi_template_as_df(self, file_path: str, sheet_name: str, header: int) -> pd.DataFrame:
        # Load FAIRe excel template as a data frame based on the specified template sheet name
        
        return pd.read_excel(file_path, sheet_name=sheet_name, header=header)
    
    def clean_samp_df(self, df:pd.DataFrame):
        # removes all FAIRe missing values and returns empty values
        
        df_clean = df.replace(self.faire_missing_values, '')
        return df_clean
    
    def get_ncbi_sample_df(self):
        # TODO: Edit elif_faire_unit col statement to just use the corresponding unit col by removing _unit from main col name. See update_unit_colums_with_no_corresponding_val method in SampleMapper to see how to do..
        
        updated_df = pd.DataFrame()

        # First handle unit transormations
        for ncbi_col_name, faire_mapping in faire_to_ncbi_units.items():
            # Get FAIRe column mame
            faire_col = faire_mapping['faire_col']

            if faire_col not in self.faire_sample_df:
                continue
            else:
                # Check if we have a constant unit or a unit column
                if 'constant_unit_val' in faire_mapping:
                    unit_val = faire_mapping['constant_unit_val']
                    updated_df[ncbi_col_name] = (self.faire_sample_df[faire_col].astype(str) + ' ' + unit_val).where(self.faire_sample_df[faire_col].astype(str).str.strip() != '', '')
                elif 'faire_unit_col' in faire_mapping:
                    unit_col = faire_mapping['faire_unit_col']
                    updated_df[ncbi_col_name] = self.faire_sample_df[faire_col].astype(str) + ' ' + self.faire_sample_df[unit_col].astype(str).where(self.faire_sample_df[faire_col].astype(str).str.strip() != '', '')
                else:
                    updated_df[ncbi_col_name] = self.faire_sample_df[faire_col]

        # Second handle direct column mappings
        for old_col_name, new_col_name in ncbi_faire_to_ncbi_column_mappings_exact.items():
            if old_col_name in self.faire_sample_df:
                updated_df[new_col_name] = self.faire_sample_df[old_col_name]

        # Third handle logic cases (depth and lat/lon)
        updated_df['*depth'] = self.faire_sample_df.apply(
                lambda row: self.get_ncbi_depth(metadata_row=row),
                axis=1
            )
        updated_df['*lat_lon'] = self.faire_sample_df.apply(
            lambda row: self.get_ncbi_lat_lon(metadata_row=row),
            axis=1
        )

        # Fourth add organism
        updated_df['*organism'] = self.ncbi_organism

        # Fifth add description for PCR technical replicates and additional column - needed to differentiate their metadata for submission to be accepted
        updated_df['description'] = self.faire_sample_df['samp_name'].apply(self.add_description_for_pcr_reps)
        updated_df['technical_rep_id'] = self.faire_sample_df['samp_name'].apply(self.add_pcr_technical_rep)

        # Get Bioproject accession from exp run df and biosample acession
        updated_df[self.ncbi_bioprojec_col_name] = updated_df[self.ncbi_samp_name_col].map(self.ncbi_bioproject_dict)
     
        # drop technical_rep_id if its empty
        if 'technical_rep_id' in updated_df.columns and updated_df['technical_rep_id'].isnull().all():
            updated_df = updated_df.drop(columns=['technical_rep_id'])

        # 6th add additional column in FAIRe sample df that do not exist in ncbi template
        final_ncbi_df = self.add_additional_faire_cols(faire_samp_df=self.faire_sample_df, updated_ncbi_df=updated_df)

        # 7th drop fields that should not be included in NCBI submission (like within_5km and rel_cont_id)
        last_final_ncbi_df = self.drop_samp_cols_not_meant_for_submission(final_ncbi_df)

        # Add biosample_accession - did not like it when I added it before this.
        last_final_ncbi_df['biosample_accession'] = last_final_ncbi_df[self.ncbi_samp_name_col].map(self.ncbi_biosample_dict)
        # drop biosample_accession if its empty (for non-osu runs)
        if 'biosample_accession' in updated_df.columns and updated_df['biosample_accession'].isnull().all():
            updated_df = updated_df.drop(columns=['biosample_accession'])
        
        return last_final_ncbi_df

    def get_sra_df(self) -> pd.DataFrame:

        updated_df = pd.DataFrame()

        # First handle direct mappings
        for old_col_name, new_col_name in ncbi_faire_sra_column_mappings_exact.items():
            if old_col_name in self.faire_experiment_run_df:
                updated_df[new_col_name] = self.faire_experiment_run_df[old_col_name]

        # Second add values that are the same across all samples
        updated_df['library_strategy'] = self.ncbi_library_strategy
        updated_df['library_source'] = self.ncbi_library_source
        updated_df['library_selection'] = self.ncbi_library_selection
        updated_df['library_layout'] = self.ncbi_library_layout
        updated_df['platform'] = self.library_prep_bebop['platform']
        updated_df['instrument_model'] = self.library_prep_bebop['instrument']
        updated_df['design_description'] = f"Sequencing performed at {self.library_prep_bebop.get('sequencing_location')}"
        updated_df['filetype'] = self.ncbi_file_type

        # # Add biosample_accession - did not like it when I added it before this.
        updated_df['biosample_accession'] =  updated_df['sample_name'].map(self.ncbi_biosample_dict)
        # drop biosample_accession if its empty (for non-osu runs)
        if 'biosample_accession' in updated_df.columns and updated_df['biosample_accession'].isnull().all():
            updated_df = updated_df.drop(columns=['biosample_accession'])
       
        # Add srr_accession - did not like it when I added it before this.
        updated_df['srr_accession'] = updated_df['library_ID'].map(self.ncbi_srr_dict)
        # drop srr_accession if its empty (for non-osu runs)
        if 'srr_accession' in updated_df.columns and updated_df['srr_accession'].isnull().all():
            updated_df = updated_df.drop(columns=['srr_accession'])

        return updated_df

    def load_beBop_yaml_terms(self, path_to_bebop: str):
        # read BeBOP yaml terms
        with open(path_to_bebop, 'r', encoding='utf-8') as f:
            post = frontmatter.load(f)
            return post
        
    def retrive_github_bebop(self, owner: str, repo: str, file_path: str):
        url = f"https://api.github.com/repos/{owner}/{repo}/contents/{file_path}"

        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()

            if 'content' in data:
                # Decode base64 to get the raw markdown file
                base64_content = data['content'].replace('\n', '').replace(' ', '')
                markdown_content = base64.b64decode(base64_content).decode('utf-8')
                
                with tempfile.NamedTemporaryFile(mode='w', suffix='.md', delete=True, encoding='utf-8') as temp_file:
                    temp_file.write(markdown_content)
                    temp_file_path = temp_file.name
                    post = self.load_beBop_yaml_terms(path_to_bebop=temp_file_path)
                    return post.metadata
        
        except requests.exceptions.RequestException as e:
            print(f"Error fetching bebop: {e}")
            return None
        
    def get_faire_cols_that_map_to_ncbi(self) -> list:
        # Creates a list of all the columns in FAIRe that map to NCBI (from the .lists)
        faire_maps_to_ncbi_cols = []

        # Add unit/and constant value faire cols
        for key, details in faire_to_ncbi_units.items():
            if "faire_col" in details:
                faire_maps_to_ncbi_cols.append(details["faire_col"])
            if "faire_unit_col" in details:
                faire_maps_to_ncbi_cols.append(details["faire_unit_col"])

        # Add exact mapping faire cols
        for key in ncbi_faire_to_ncbi_column_mappings_exact.keys():
            faire_maps_to_ncbi_cols.append(key)

        return faire_maps_to_ncbi_cols
    
    def get_additional_faire_unit_cols_dict(self, faire_cols: list) -> dict:
        # Creates a dictionary of the additional faire_cols with that have corresponding unit cols and creates a dict with faire col as key and unit col as value
        unit_col_dict = {}
        for col in faire_cols:
            unit_col = f"{col}_unit"
            units_col = f"{col}_units"
            if unit_col in faire_cols:
                unit_col_dict[col] = unit_col
            if units_col in faire_cols:
                unit_col_dict[col] = units_col

        # remove columns from list that exist in dictionary
        to_remove = set(unit_col_dict.keys()) 
        to_remove.update(unit_col_dict.values())
        filtered_faire_list = [col for col in faire_cols if col not in to_remove]
        
        return unit_col_dict, filtered_faire_list

    def add_additional_faire_cols(self, faire_samp_df: pd.DataFrame, updated_ncbi_df: pd.DataFrame) -> pd.DataFrame:
        # Adds additional FAIRe columns not in the NCBI sample template
       
        # Get unit cols and regular cols - additional_reg_faire_cols is a list of regular columsn (not unit corresponding ones)
        # And additional_cols_units_dict is columns with unit columsn
        faire_cols_that_map = set(self.get_faire_cols_that_map_to_ncbi())
        faire_cols = set(faire_samp_df.columns)
        additional_faire_cols = list(faire_cols - faire_cols_that_map)
        cols = self.get_additional_faire_unit_cols_dict(faire_cols=additional_faire_cols)
        additional_cols_units_dict = cols[0]
        additional_reg_faire_cols = cols[1]

        # Add unit cols first
        for key, value in additional_cols_units_dict.items():
            updated_ncbi_df[key] = faire_samp_df[key].astype(str) + ' ' + faire_samp_df[value].astype(str)

        # Then add regular columns (those with no corresponding units)
        for faire_col in additional_reg_faire_cols:
            updated_ncbi_df[faire_col] = faire_samp_df[faire_col]

        # Drop columns that are empty for all rows that were additional faire columns
        columns_to_drop = []
        for col in additional_faire_cols:
            if col in updated_ncbi_df.columns:
                if updated_ncbi_df[col].isna().all() or updated_ncbi_df[col].astype(str).str.strip().eq('').all():
                    columns_to_drop.append(col)
        final_ncbi_df = updated_ncbi_df.drop(columns=columns_to_drop)

        return final_ncbi_df

    def get_ncbi_depth(self, metadata_row: pd.Series) -> str:
        # Get the ncbi formatted value for depth, which is the interval of minimumDepthInMeters - maximumDepthInMeters
        min_depth = metadata_row['minimumDepthInMeters']
        max_depth = metadata_row['maximumDepthInMeters']

        # If min_depth and max_depth are not the same, report as interval
        if min_depth != max_depth and max_depth != '':
            ncbi_depth = min_depth + ' ' + 'm' + ' - ' + max_depth + ' ' + 'm'
        # if min_depth and max_depth are the same, report just one because they will be the same
        elif min_depth == max_depth and max_depth != '':
            ncbi_depth = min_depth + ' ' + 'm'
        else:
            if min_depth == '' and max_depth == '': # for controll samples
                return ''
            else:
                raise ValueError(f"Something wrong with the depth. Min depth is {min_depth} and max_depth is {max_depth}. {metadata_row}")

        return ncbi_depth

    def get_ncbi_lat_lon(self, metadata_row: pd.Series) -> str:
        try:
            lat = float(metadata_row['decimalLatitude'])
            lon = float(metadata_row['decimalLongitude'])

            lat_dir = 'N' if lat >= 0 else 'S'
            lon_dir = 'E' if lon >= 0 else 'W'

            # Take absolute values
            lat_abs = abs(lat)
            lon_abs = abs (lon)

            lat_formatted = f"{lat_abs:.4f} {lat_dir}"
            lon_formatted = f"{lon_abs:.4f} {lon_dir}"

            ncbi_lat_lon = f"{lat_formatted} {lon_formatted}"

            return ncbi_lat_lon
        except:
            if metadata_row['decimalLatitude'] == '' and metadata_row['decimalLongitude'] == '':
                return ''
            else:
                raise ValueError(f"Something wrong with lat lon and can't transform for sample: {metadata_row[self.faire_samp_name_col]} with lat/lon {metadata_row['decimalLatitude']}/{metadata_row['decimalLongitude']}")
    
    def add_description_for_pcr_reps(self, sample_name: str) -> str:
        # Add a description for the PCR technical replicates (they have to be distinguishable to submit)
        if 'PCR' in sample_name:
            samp_parts = sample_name.split('.')
            # PCR number is last character of the last part
            pcr_num = samp_parts[-1][-1]
            original_sample = '.'.join(samp_parts[:-1])
            return f"PCR technical replicate number {pcr_num} of {original_sample}"
        else:
            return ''

    def add_pcr_technical_rep(self, sample_name: str) -> str:
        if 'PCR' in sample_name:
            samp_parts = sample_name.split('.')
            # PCR number is last character of the last part
            pcr_num = samp_parts[-1][-1]
            return pcr_num
        else:
            return ''
    
    def save_to_excel_template(self, template_path: str, ncbi_excel_save_path: str, sheet_name: str, header: int, final_ncbi_df: pd.DataFrame):

        # Copy template file at saved locationsample_ncbi_excel_save_path
        shutil.copy2(template_path, ncbi_excel_save_path)

        # Load the template to get the column structure
        template_df = pd.read_excel(template_path, sheet_name=sheet_name, header=header)
        template_columns = template_df.columns.tolist()
        
        # Load workbook to preserve formatting
        book = load_workbook(ncbi_excel_save_path)
        ws = book[sheet_name]

        # find column in final_ncbi_df that don't exist in template df (e.g. technical replicate)
        missing_cols = [col for col in final_ncbi_df.columns if col not in template_columns]

        # Add missing columns
        if missing_cols:
            next_col_idx = len(template_columns) + 1
            for i, col_name in enumerate(missing_cols):
                # Add header for new column
                ws.cell(row=header+1, column=next_col_idx + i, value=col_name)
                template_columns.append(col_name)

        for col_idx, template_col_name in enumerate(template_columns, start=1):
            if template_col_name in final_ncbi_df.columns:
                #write data for this column
                column_data = final_ncbi_df[template_col_name]
                for row_idx, value in enumerate(column_data, start = header+2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
        # save workbook
        book.save(ncbi_excel_save_path)

        # # Save as tsv
        # df = pd.read_excel(sample_ncbi_excel_save_path, sheet_name=sheet_name)
        # df.to_csv(sample_ncbi_excel_save_path.replace('.xlsx', '.tsv'), sep='\t', index=False)

        print(f'NCBI sample saved to {ncbi_excel_save_path}')

    def drop_samp_cols_not_meant_for_submission(self, sample_df: pd.DataFrame) -> pd.DataFrame:
        # Drop columns not meant for submission like stations_within_5m
        if self.faire_stations_in_5km_col in sample_df.columns:
            return sample_df.drop(columns=[self.faire_stations_in_5km_col])
        else:
            return sample_df

    def get_ncbi_bioproject_if_exists(self, metadata_row: pd.Series, id_prefix: str) -> str:
        # Uses the associatedSequences column in the FAIRe df to get the NCBI accession number 
        # # (e.g. for bioproject, srr, and biosample):
        try:
            associated_sequences = metadata_row['associatedSequences']
            associated_seqs = associated_sequences.split(' | ')
            for accession_num in associated_seqs:
                if id_prefix in accession_num:
                    match = re.search(rf'{id_prefix}\w+', accession_num)
                    if match:
                        ncbi_accession_id = match.group()
                        return ncbi_accession_id
        except:
            pass

    def create_ncbi_accession_dict_project_and_biosample(self, id_prefix: str) -> dict:
        # Creates a dictionary of accessions for ncbi (bioproject, biosample) 
        # Don't use for SRA because some samples will have duplicate or more samp_names and the values will get overwritten. This is
        # only for PRJNA and SAMN
        if id_prefix == 'SRR':
           group_col = 'lib_id'
        else:
            group_col = self.faire_samp_name_col

        exp_df = self.faire_experiment_run_df.copy()
        exp_df[id_prefix] = exp_df.apply(
            lambda row: self.get_ncbi_bioproject_if_exists(metadata_row=row, id_prefix=id_prefix),
            axis=1 
        )

        # Handle duplicate samp_name values and prioritize non-Non accessions
        grouped = exp_df.groupby(group_col)[id_prefix]

        # Use a dictionary comprehension to find the first non-None value for each group
        if group_col is self.faire_samp_name_col:
            ncbi_accession_dict = {
                samp_name: group.dropna().iloc[0] if not group.dropna().empty else None
                for samp_name, group in grouped
            }
        elif group_col is 'lib_id':
            ncbi_accession_dict = dict(zip(exp_df[group_col], exp_df[id_prefix]))
        
        return ncbi_accession_dict


 
