import pandas as pd
import openpyxl
import copy
from openpyxl.styles import PatternFill
from .faire_mapper import OmeFaireMapper

# TODO: change 'config file' not in value in fill_out_analysis_metadata() method when BeBOP is finished an we know what it will say to reference config file

class AnalysisMetadataMapper(OmeFaireMapper):

    faire_template_analysis_sheet_name = 'analysisMetadata'
    project_id_col = 'project_id'
    assay_name_col = 'assay_name'
    analysis_run_name_col = 'analysis_run_name'

    def __init__(self, config_yaml, bioinformatics_bebop_path: str, bioinformatics_config_google_sheet_id: str, experiment_run_metadata_df: pd.DataFrame,
                 project_id: str, bioinformatics_software_name: str, bebop_config_run_col_name: str, bebop_config_marker_col_name: str):
        
        super().__init__(config_yaml)

        self.bioiformatics_bebop = self.load_beBop_yaml_terms(path_to_bebop=bioinformatics_bebop_path)
        self.bioinformatics_config_df = self.load_google_sheet_as_df(google_sheet_id=bioinformatics_config_google_sheet_id, sheet_name='Sheet1', header=0)
        self.experiment_run_df = experiment_run_metadata_df
        self.bioinformatics_software_name = bioinformatics_software_name
        self.project_id = project_id
        self.bebop_config_run_col_name = bebop_config_run_col_name
        self.bebop_config_marker_col_name = bebop_config_marker_col_name
        self.analysis_run_dict = self.create_analysis_run_dict()

    def process_analysis_metadata(self):
        analysis_metadata_df = self.fill_out_analysis_metadata()
        self.save_to_excel(final_analysis_metadata_df=analysis_metadata_df, excel_file_to_save_to=self.final_faire_template_path)

    def load_analyis_metadata_df(self):
        df = self.load_faire_template_as_df(file_path=self.faire_template_file,sheet_name=self.faire_template_analysis_sheet_name, header=0)
        cols = df['term_name'].tolist()
        empty_analysis_metadata_df = pd.DataFrame(columns=cols)
        
        return empty_analysis_metadata_df
    
    def create_analysis_run_dict(self):
        # creates a dictionary of all unique analysis run name (E.g. REVAMP_Parada16S_osu876) as the key and the assay name as the value
        df = self.experiment_run_df.copy()
        df['analysis_run_name'] = self.bioinformatics_software_name + '_' + df['lib_id'].apply(lambda x: '_'.join(x.split('_')[-2:]))
        analysis_run_dict = df.drop_duplicates(['analysis_run_name', 'assay_name']).set_index('analysis_run_name')['assay_name'].to_dict()
        return analysis_run_dict

    def query_config_file_by_run_and_marker_for_attribute(self, marker: str, run_name: str, faire_attribute: str) -> pd.DataFrame:
        # Searches config file by marker and run name and faire attribute to return the value
        matching_cols = [col for col in self.bioinformatics_config_df.columns if faire_attribute in col]
        
        # Cast columns in lower case
        self.bioinformatics_config_df[self.bebop_config_run_col_name] = self.bioinformatics_config_df[self.bebop_config_run_col_name].str.lower()
        self.bioinformatics_config_df[self.bebop_config_marker_col_name] = self.bioinformatics_config_df[self.bebop_config_marker_col_name].str.lower()
        
        query_str = f"{self.bebop_config_run_col_name} == @run_name & {self.bebop_config_marker_col_name} == @marker"
        filtered_df = self.bioinformatics_config_df.query(query_str)

        result_parts = []
        for col in matching_cols:
            # Get the first part of the column name (before first semicolon)
            first_part = col.split(';')[0]

            # Get the column value from the filtered row
            col_value = filtered_df[col].iloc[0]

            # Format as "first_part: val"
            result_parts.append(f"{first_part}: {col_value}")
        
        result = " | ".join(result_parts)
        return result
        
        return value

    def fill_out_analysis_metadata(self):

        template_df = self.load_analyis_metadata_df()
        rows_data = []

        for analysis_run_name, assay_name in self.analysis_run_dict.items():
            # First handle values from the BeBOP that are the same across analyses
            row_dict = {
                self.project_id_col: self.project_id,
                self.assay_name_col: assay_name,
                self.analysis_run_name_col: analysis_run_name
            }
            
            # If faire_attribute from BeBOP exists as column name in df, then add value
            for faire_attribute, value in self.bioiformatics_bebop.metadata.items():
                if value is not None:
                    if faire_attribute in template_df.columns and 'config file' not in str(value):
                        row_dict[faire_attribute] = value
                    elif faire_attribute in template_df.columns and 'config file' in value:
                        marker = analysis_run_name.split('_')[1].lower()
                        run_name = analysis_run_name.split('_')[2].lower()
                        try:
                            row_dict[faire_attribute] = self.query_config_file_by_run_and_marker_for_attribute(marker=marker, run_name=run_name, faire_attribute=faire_attribute)
                            
                        except:
                            print(f"faire_attribute {faire_attribute} does not seem to exist in config file, though it says 'config file' in bebop")
                rows_data.append(row_dict)
       
        final_analysis_df = pd.DataFrame(rows_data)
        
        return final_analysis_df
        
    def save_to_excel(self, final_analysis_metadata_df: pd.DataFrame, excel_file_to_save_to: str):
        # Save analysisMetadata df rows to their own excel file sheets
        
        # for formatting
        template_wb = openpyxl.load_workbook(self.faire_template_file)
        template_sheet = template_wb[self.faire_template_analysis_sheet_name] 
        
        # Create new workbook
        new_wb = openpyxl.load_workbook(excel_file_to_save_to)
        
        for idx, data_row in final_analysis_metadata_df.iterrows():
            sheet_name = f'analysisMetadata_{data_row[self.analysis_run_name_col]}'

            new_sheet = new_wb.create_sheet(title=sheet_name)

            # Copy all cells and formatting from template
            for row in template_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column = cell.column)
                    new_cell.value = cell.value

                    # Copy formatting
                    if cell.fill:
                        new_cell.fill = copy.copy(cell.fill)
                    if cell.font:
                        new_cell.font = copy.copy(cell.font)
                    if cell.alignment:
                        new_cell.alignment = copy.copy(cell.alignment)
                    if cell.border:
                        new_cell.border = copy.copy(cell.border)
                    if cell.number_format:
                        new_cell.number_format = copy.copy(cell.number_format)

            for row in new_sheet.iter_rows(min_row=2):
                #skip header row
                term_name_cell = row[2]
                values_cell = row[3]
                
                if term_name_cell.value and term_name_cell.value in final_analysis_metadata_df.columns:
                    values_cell.value = data_row[term_name_cell.value]

            new_wb.save(excel_file_to_save_to)
            template_wb.close()
